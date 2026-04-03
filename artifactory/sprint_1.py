#!/usr/bin/python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.drawing.image import Image
from configparser import ConfigParser
import requests
import concurrent.futures
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import matplotlib.pyplot as plt
import numpy as np
import os

# ======== 全局配置 ========
MAX_WORKERS = 10  # 并发线程数
TIMEOUT = 10      # 请求超时秒数

# ======== 网络会话，支持重试 ========
session = requests.Session()
retries = Retry(total=3, backoff_factor=0.5, status_forcelist=[500, 502, 503, 504])
session.mount("https://", HTTPAdapter(max_retries=retries))
session.mount("http://", HTTPAdapter(max_retries=retries))

# ======== 获取 JIRA issues 列表 ========
def get_jira_issues(jql, JIRA_URL, API_TOKEN):
    all_issues = []
    start_at = 0
    max_results = 100

    while True:
        url = f"{JIRA_URL}/rest/api/2/search"
        headers = {"Authorization": f"Bearer {API_TOKEN}", "Content-Type": "application/json"}
        params = {
            "jql": jql,
            "startAt": start_at,
            "maxResults": max_results,
            "fields": "key,summary,status,assignee,priority,updated"
        }

        response = session.get(url, headers=headers, params=params, timeout=TIMEOUT)
        if response.status_code != 200:
            print(f"❌❌ JIRA 请求失败: {response.status_code}")
            break

        data = response.json()
        issues = data.get("issues", [])
        all_issues.extend(issues)
        if len(issues) < max_results:
            break
        start_at += max_results

    return all_issues

# ======== 提取测试状态 ========
def extract_test_status_info(data):
    passed = total = 0

    def find_customfield(obj):
        if isinstance(obj, dict):
            if "customfield_11120" in obj:
                return obj["customfield_11120"]
            for v in obj.values():
                result = find_customfield(v)
                if result:
                    return result
        elif isinstance(obj, list):
            for item in obj:
                result = find_customfield(item)
                if result:
                    return result
        return None

    custome = find_customfield(data)
    if not custome:
        return None

    total = custome.get("count", 0)
    for status in custome.get("statuses", []):
        if status.get("name") == "PASS":
            passed = status.get("statusCount", 0)

    return {"total": total, "passed": passed}

# ======== 并发获取 JIRA issue 详情 ========
def get_issue_detail(issue, JIRA_URL, API_TOKEN):
    jira_key = issue["key"]
    summary = issue["fields"]["summary"]
    version_type = summary.split("_")[0].strip()
    if version_type not in ("STAR3.0", "STAR3.5"):
        return None

    url = f"{JIRA_URL}/rest/api/2/issue/{jira_key}"
    headers = {"Authorization": f"Bearer {API_TOKEN}"}
    try:
        response = session.get(url, headers=headers, timeout=TIMEOUT)
        response.raise_for_status()
        data = response.json()
        return jira_key, extract_test_status_info(data)
    except Exception as e:
        print(f"⚠️ 获取 {jira_key} 失败: {e}")
        return None

def get_tmx_info_parallel(issues, JIRA_URL, API_TOKEN):
    print("🚀🚀 正在并发获取 JIRA issue 详情 ...")
    all_tmx_info = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(get_issue_detail, issue, JIRA_URL, API_TOKEN) for issue in issues]
        for future in concurrent.futures.as_completed(futures):
            result = future.result()
            if result:
                key, info = result
                if info:
                    all_tmx_info[key] = info
    return all_tmx_info

# ======== GitLab MR 查询（带缓存） ========
def query_mr(mr_label, GITLAB_URL, GROUP_PATH, ACCESS_TOKEN, mr_cache):
    if mr_label in mr_cache:
        return mr_cache[mr_label]

    new_label = f"integrated_{mr_label}"
    url = f"{GITLAB_URL}/api/v4/groups/{GROUP_PATH}/merge_requests"
    headers = {"PRIVATE-TOKEN": ACCESS_TOKEN}
    params = {"scope": "all", "per_page": 100, "labels": new_label}

    try:
        response = session.get(url, params=params, headers=headers, timeout=TIMEOUT)
        if response.status_code == 200:
            mrs = response.json()
            mr_count = len(mrs)
            mr_cache[mr_label] = mr_count
            return mr_count
        else:
            print(f"⚠️ GitLab 请求失败 {response.status_code}")
            return 0
    except Exception as e:
        print(f"⚠️ MR 查询失败: {e}")
        return 0

# ======== 写入 Excel ========
def write_excel(issues, tmx_info, GITLAB_URL, GROUP_PATH, ACCESS_TOKEN, file_path="rsu_staging_trend.xlsx"):
    wb = Workbook()
    # 移除默认创建的空工作表
    if wb.active:
        wb.remove(wb.active)
    
    mr_cache = {}
    summary_data = {}
    all_summary_rows = []

    # 处理每个issue
    for issue in issues:
        key = issue["key"]
        if key not in tmx_info:
            continue

        summary = issue["fields"]["summary"].strip()
        parts = summary.split("_")
        if len(parts) < 3:
            continue

        version_type = parts[0].strip()
        build_type = "_".join(parts[1:3])
        if not (build_type.endswith("_DEV") or build_type.endswith("_HQ")):
            continue

        version_base = parts[1]
        version_label = f"{version_base}_DEV"
        config_key = f"{version_type}_{version_base.split('-')[0]}"

        summary_data.setdefault(config_key, {})
        summary_data[config_key].setdefault(version_label, {
            "passed": 0, "total": 0, "mr_merged": 0, "mr_submitted": 0, "mr_reviewed": 0
        })

        mr_label = parts[1]
        mr_number = query_mr(mr_label, GITLAB_URL, GROUP_PATH, ACCESS_TOKEN, mr_cache)
        data = tmx_info[key]
        
        summary_data[config_key][version_label]["passed"] += data.get("passed", 0)
        summary_data[config_key][version_label]["total"] += data.get("total", 0)
        summary_data[config_key][version_label]["mr_merged"] = mr_number
        # 为演示添加模拟数据（保持与原始代码一致）
        summary_data[config_key][version_label]["mr_submitted"] = mr_number + 2
        summary_data[config_key][version_label]["mr_reviewed"] = mr_number + 1

    # 设置单元格样式（保持原始样式）
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    align_center = Alignment(horizontal="center", vertical="center")

    # === 写 Excel ===（保持原始格式）
    for config_key, versions in summary_data.items():
        ws = wb.create_sheet(config_key[:31])
        ws.append([config_key])
        ws.append(["Passed"])
        ws.append(["Total"])
        ws.append(["MR Submitted"])
        ws.append(["MR Reviewed"])
        ws.append(["MR merged"])
        ws.append(["Passed rate"])

        col = 2
        for version_label, vals in versions.items():
            ws.cell(row=1, column=col, value=version_label)
            ws.cell(row=2, column=col, value=vals["passed"])
            ws.cell(row=3, column=col, value=vals["total"])
            ws.cell(row=4, column=col, value=vals["mr_merged"]+2)
            ws.cell(row=5, column=col, value=vals["mr_merged"]+1)
            ws.cell(row=6, column=col, value=vals["mr_merged"])

            total_val = vals["total"]
            rate = f"{(vals['passed'] / total_val) * 100:.2f}%" if total_val > 0 else "0%"
            ws.cell(row=7, column=col, value=rate)
            col += 1

        # 自动调整列宽 + 边框 + 居中（保持原始样式）
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                cell.border = thin_border
                cell.alignment = align_center
            ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2

        # === 绘制图表 ===（保持原始图表格式样式）
        records = []
        for version_label, vals in versions.items():
            records.append({
                "build": version_label,
                "passed": vals["passed"],
                "total": vals["total"],
                "mr_submitted": vals["mr_submitted"],
                "mr_reviewed": vals["mr_reviewed"],
                "mr_merged": vals["mr_merged"],
            })

        if not records:
            continue

        builds = [r["build"].replace("_DEV", "") for r in records]
        passed_cases = [r["passed"] for r in records]
        total_cases = [r["total"] for r in records]
        failed_cases = [t - p for t, p in zip(total_cases, passed_cases)]
        mr_sub = [r["mr_submitted"] for r in records]
        mr_rev = [r["mr_reviewed"] for r in records]
        mr_mer = [r["mr_merged"] for r in records]
        rates = [(p / t * 100 if t else 0) for p, t in zip(passed_cases, total_cases)]

        # 保持原始图表计算逻辑
        bar_width = 0.2
        gap = 0.2
        n = len(builds)
        if n == 1:
            x = np.array([0.5])
            fig_width = 2.5
        else:
            x = np.arange(n) * (bar_width + gap)
            fig_width = 2 + n * (bar_width + gap) * 2

        fig_height = 6
        xtick_fontsize = 8 if n <= 20 else max(6, int(160 / n))

        # 创建图表（保持原始样式）
        fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=200)
        fig.patch.set_facecolor("#2b2b2b")
        ax.set_facecolor("#2b2b2b")

        # 绘制柱状图和折线图（保持原始颜色和样式）
        ax.bar(x, passed_cases, width=bar_width, color="#4CAF50", edgecolor="white", linewidth=0.7, label="Passed")
        ax.bar(x, failed_cases, width=bar_width, bottom=passed_cases, color="#2196F3", edgecolor="white", linewidth=0.7, label="Total")
        ax.plot(x, mr_sub, color="#FFC107", linewidth=2, marker="o", label="MR Submitted")
        ax.plot(x, mr_rev, color="#9C27B0", linewidth=2, marker="o", label="MR Reviewed")
        ax.plot(x, mr_mer, color="#F44336", linewidth=2, marker="o", label="MR Merged")

        # 添加通过率标签（保持原始样式）
        max_total = max(total_cases) if total_cases else 0
        for i, r in enumerate(rates):
            ax.text(x[i], total_cases[i] + max_total * 0.03, f"{r:.1f}%", 
                   ha="center", va="bottom", fontsize=5, color="white", fontweight="bold")

        # 设置坐标轴（保持原始样式）
        ax.set_xticks(x)
        ax.set_xticklabels(builds, fontsize=xtick_fontsize, color="white")
        ax.tick_params(axis="y", colors="white")
        ax.grid(axis="y", linestyle="--", alpha=0.3, color="white")
        ax.set_title(f"RSU {config_key} DEV MR Test", fontsize=7, fontweight="bold", color="white", pad=12)
        ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.08), ncol=5, frameon=False, fontsize=4, labelcolor="white")

        plt.tight_layout(pad=2.0)

        # 保存图表
        chart_dir = "charts"
        if not os.path.exists(chart_dir):
            os.makedirs(chart_dir)
        chart_path = os.path.join(chart_dir, f"{config_key}_chart.png")
        plt.savefig(chart_path, bbox_inches="tight", facecolor=fig.get_facecolor(), dpi=200)
        plt.close()

        # 插入图表到Excel（保持原始位置）
        if os.path.exists(chart_path):
            try:
                img = Image(chart_path)
                img.width, img.height = int(fig_width * 120), int(fig_height * 120)
                ws.add_image(img, "A10")
                print(f"✅ 图表已添加到工作表: {config_key}")
            except Exception as e:
                print(f"⚠️ 添加图表失败: {e}")

        # 为Summary页准备数据
        for version_label, vals in versions.items():
            total = vals.get("total", 0)
            passed = vals.get("passed", 0)
            pass_rate = f"{(passed / total) * 100:.2f}%" if total > 0 else "0%"
            
            all_summary_rows.append({
                "sheet": config_key,
                "build": version_label,
                "passed": passed,
                "total": total,
                "passed_rate": pass_rate,
                "mr_submitted": vals.get("mr_submitted", 0),
                "mr_reviewed": vals.get("mr_reviewed", 0),
                "mr_merged": vals.get("mr_merged", 0)
            })

    # === 生成 Summary 总览页 ===（保持原始格式）
    summary_name = "Summary"
    if summary_name in wb.sheetnames:
        ws_summary = wb[summary_name]
        wb.remove(ws_summary)
    ws_summary = wb.create_sheet(summary_name, 0)

    headers = ["Sheet", "Build", "Passed", "Total", "Passed rate", "MR submitted", "MR reviewed", "MR merged"]
    ws_summary.append(headers)

    for row in all_summary_rows:
        ws_summary.append([
            row["sheet"], row["build"], row["passed"], row["total"],
            row["passed_rate"], row["mr_submitted"], row["mr_reviewed"], row["mr_merged"]
        ])

    # 自动列宽（保持原始样式）
    for column in ws_summary.columns:
        max_len = max(len(str(c.value or "")) for c in column)
        ws_summary.column_dimensions[column[0].column_letter].width = (max_len + 2) * 1.2

    # 保存文件
    wb.save(file_path)
    print(f"✅ Excel 文件生成/更新完成: {file_path}")
    print(f"📊📊 Summary 汇总页已更新，共 {len(all_summary_rows)} 条记录。")

# ======== 主函数 ========
if __name__ == "__main__":
    # 读取配置
    config = ConfigParser()
    config.read("/home/gaoyuxia/config.ini")
    
    ACCESS_TOKEN = config.get("gitlab", "token")
    GITLAB_URL = config.get("gitlab", "url")
    GROUP_PATH = "RSE"
    JIRA_URL = config.get("jira", "JIRA_URL")
    API_TOKEN = config.get("jira", "API_TOKEN")
    
    # JQL查询
    jql_query = 'project = TMX AND issuetype = "Test Execution" AND reporter in (shaoli) ORDER BY priority DESC, updated DESC'
    
    # 获取issues
    issues = get_jira_issues(jql_query, JIRA_URL, API_TOKEN)
    print(f"📋📋 共找到 {len(issues)} 个 issues")
    
    if issues:
        # 获取详细信息
        tmx_info = get_tmx_info_parallel(issues, JIRA_URL, API_TOKEN)
        print(f"✅ 成功获取 {len(tmx_info)} 个issue的详细信息")
        
        # 生成Excel和图表
        write_excel(issues, tmx_info, GITLAB_URL, GROUP_PATH, ACCESS_TOKEN)
    else:
        print("❌ 未找到符合条件的issues")