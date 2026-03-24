# 得到版本信息从swfcn

from configparser import ConfigParser
import subprocess
from pathlib import Path
import json
from typing import Dict, List, Optional
from datetime import datetime
from collections import defaultdict
import requests
import sys
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import Alignment


def query_swf_artifactory(ARTIFACTORY_URL_SWF,ARTIFACTORY_TOKEN_SWF,SEARCH_PATH):
    """查询Artifactory并保存结果到path.txt"""
    cmd = [
        "jf", "rt", "search",
        SEARCH_PATH,
        "--url", ARTIFACTORY_URL_SWF,
        "--user", "GAYUXIA",
        "--password", ARTIFACTORY_TOKEN_SWF,
        "--insecure-tls=false",
        "--recursive"
    ]
    
    try:
        print("正在查询Artifactory...")
        result = subprocess.run(
        cmd,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE
        )
        Path("version_swf.txt").write_text(result.stdout, encoding="utf-8")
        print("查询结果已保存到 version_swf.txt")
        return True
    except subprocess.CalledProcessError as e:
        print(f"查询失败: {e.stderr}")
        raise


def query_swfcn_artifactory(ARTIFACTORY_URL_SWFCN,ARTIFACTORY_TOKEN_SWFCN,SEARCH_PATH):
    """查询Artifactory并保存结果到path.txt"""
    cmd = [
        "jf", "rt", "search",
        SEARCH_PATH,
        "--url", ARTIFACTORY_URL_SWFCN,
        "--user", "GAYUXIA",
        "--password", ARTIFACTORY_TOKEN_SWFCN,
        "--insecure-tls=false",
        "--recursive"
    ]
    
    try:
        print("正在查询Artifactory...")
        result = subprocess.run(
        cmd,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE
        )
        Path("version_swfcn.txt").write_text(result.stdout, encoding="utf-8")
        print("查询结果已保存到 version_swfcn.txt")
        return True
    except subprocess.CalledProcessError as e:
        print(f"查询失败: {e.stderr}")
        raise


def parse_build_paths(version_file, artifact_url):
    # 读取文件
    with open(version_file, "r") as f:
        data = json.load(f)

    # 初始化结果结构：{ STAR version → build_type → category → set() }
    result = defaultdict(lambda: defaultdict(lambda: {"flat_build": set(), "swup": set()}))

    def classify_path(path):
        # 分类类别：flat_build 或 swup
        if "flat_build" in path and path.endswith(".tar.gz"):
            category = "flat_build"
        elif "/swup/" in path:
            category = "swup"
        else:
            return None, None, None

        # 构建类型分类
        if "dev_userdebug" in path or "userdebug" in path:
            build_type = "dev_userdebug"
        elif "dev" in path or "dev_user" in path:
            build_type = "dev_user"
        elif "prod-pl_user" in path or "prod-pl" in path:
            build_type = "prod-pl_user"
        elif "prod-rl_user" in path or "prod-rl" in path:
            build_type = "prod-rl_user"
        else:
            return None, None, None

        # 提取 STAR 版本（如 STAR3.0 或 STAR3.5）
        if "STAR3.0" in path:
            star_version = "STAR3.0"
        elif "STAR3.5" in path:
            star_version = "STAR3.5"
        else:
            return None, None, None

        return star_version, build_type, category

    # 遍历并分类
    for entry in data:
        path = entry.get("path", "")
        star_version, build_type, category = classify_path(path)
        if star_version and build_type and category:
            if category == "swup":
                path = "/".join(path.split("/")[:11])  # 裁剪 swup 路径
            result[star_version][build_type][category].add(path)

    # 转换为 JSON 可序列化结构（set → list）
    clean_result = {
        star: {
            build: {
                kind: sorted([f"{artifact_url}/{p}" for p in paths])
                for kind, paths in builds.items()
            }
            for build, builds in star_data.items()
        }
        for star, star_data in result.items()
    }
    return clean_result


def write_excel(swfcn_clean_result, swf_clean_result, filename, ESTAND):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Build Paths Classification"

        headers = [
            "Branch/Build Date",
            "Version/Pipeline",
            "STAR Version",
            "Build Type",
            "Archive Name",
            "Archive SWFCN URL",
            "Archive SWF URL"
        ]
        ws.append(headers)

        start_row = ws.max_row + 1  # = 2

        # ================= 写数据 =================
        for star_version, build_types in swfcn_clean_result.items():

            star_has_data = False  # 记录这个 STAR 是否写过数据

            for build_type, categories in build_types.items():
                for category, swfcn_urls in categories.items():
                    swf_urls = swf_clean_result \
                        .get(star_version, {}) \
                        .get(build_type, {}) \
                        .get(category, [])

                    for idx, swfcn_url in enumerate(swfcn_urls):
                        star_has_data = True
                        swf_url = swf_urls[idx] if idx < len(swf_urls) else ""
                        ws.append([
                            "8295_master",
                            ESTAND,
                            star_version,
                            build_type,
                            category,
                            swfcn_url,
                            swf_url
                        ])

            # ⭐⭐⭐ 在该 STAR 的最后一行追加 vcpu
            if star_version in ("STAR3.0", "STAR3.5") and star_has_data:
                ws.append([
                    "8295_master",
                    ESTAND,
                    star_version,   # ⚠ 必须还是同一个 STAR
                    "vcpu",
                    "",
                    "",
                    ""
                ])

        end_row = ws.max_row

        # ================= 合并单元格 =================
        if start_row <= end_row:
            ws.merge_cells(start_row=start_row, start_column=1,
                           end_row=end_row, end_column=1)
            ws.cell(row=start_row, column=1).alignment = Alignment(vertical='center')

            ws.merge_cells(start_row=start_row, start_column=2,
                           end_row=end_row, end_column=2)
            ws.cell(row=start_row, column=2).alignment = Alignment(vertical='center')

        merge_same_cells(ws, col=3, start_row=2)  # STAR Version
        merge_same_cells(ws, col=4, start_row=2)  # Build Type
        merge_vcpu_row_columns(ws)   # ⭐ 最后再调用

        # ================= 自动列宽 =================
        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        wb.save(filename)
        print("\033[1;35m 分类结果已写入Excel文件:\033[0m", filename)

    except Exception as e:
        print(f"\033[1;31m 写入Excel文件时出错: {e} \033[0m")

def merge_same_cells(ws, col, start_row):
    max_row = ws.max_row
    current_value = None
    merge_start = start_row

    for row in range(start_row, max_row + 1):
        value = ws.cell(row=row, column=col).value

        if current_value is None:
            current_value = value
            merge_start = row
        elif value != current_value:
            if merge_start < row - 1:
                ws.merge_cells(start_row=merge_start, start_column=col,
                               end_row=row - 1, end_column=col)
                ws.cell(row=merge_start, column=col).alignment = Alignment(vertical='center')
            current_value = value
            merge_start = row

    # 处理最后一段
    if merge_start < max_row:
        ws.merge_cells(start_row=merge_start, start_column=col,
                       end_row=max_row, end_column=col)
        ws.cell(row=merge_start, column=col).alignment = Alignment(vertical='center')

def merge_vcpu_row_columns(ws):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=4).value == "vcpu":
            ws.merge_cells(start_row=row, start_column=4,
                           end_row=row, end_column=5)
            ws.cell(row=row, column=4).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("\033[1;31m 需要输入一个参数 \033[0m") 
        print("\033[1;31m 带有vcpu的行 \033[0m")
        print("\033[1;31m 版本路径swfcn：sb/sb/sb/20250722/E233.0/ \033[0m") 
        print("\033[1;31m 版本路径swf：sb/sb/sb/20250911/E237.0/ \033[0m")
        print("\033[1;31m ESTAND: E244.0 \033[0m")  
        sys.exit(1)  # 非零退出码表示错

    # 加载环境变量
    config = ConfigParser()
    config.read("/home/gayuxia/config.ini")
    ARTIFACTORY_URL_SWFCN = config.get("artifact","url")
    ARTIFACTORY_TOKEN_SWFCN = config.get("artifact","token")
    ARTIFACTORY_URL_SWF = config.get("artifact_swf","url")
    ARTIFACTORY_TOKEN_SWF = config.get("artifact_swf","token")

    patterns = ["*.txt", "*.json", "*.xlsx"]
    for pattern in patterns:
        for file_path in glob.glob(pattern):
            os.remove(file_path)


    SEARCH_PATH1 = sys.argv[1]
    SEARCH_PATH2 = sys.argv[2]
    ESTAND = sys.argv[3]
    # 从 artifactory 中查询版本信息
    query_swfcn_artifactory(ARTIFACTORY_URL_SWFCN,ARTIFACTORY_TOKEN_SWFCN,SEARCH_PATH1)
    query_swf_artifactory(ARTIFACTORY_URL_SWF,ARTIFACTORY_TOKEN_SWF,SEARCH_PATH2)
    # 解析文件，得到需要的数据信息
    swfcn_clean_result = parse_build_paths("version_swfcn.txt", ARTIFACTORY_URL_SWFCN)
    swf_clean_result = parse_build_paths("version_swf.txt", ARTIFACTORY_URL_SWF)

    # 写入 excel 表格中
    filename = "output.xlsx"
    write_excel(swfcn_clean_result, swf_clean_result ,filename, ESTAND)
 

    

