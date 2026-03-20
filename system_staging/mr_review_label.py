import requests
import json
import sys
import re
from urllib.parse import quote
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from urllib.parse import quote
from common.Query import Query




# === 配置信息 ===
GITLAB_URL = "https://git.sb.com"
GROUP_PATH = "sb"
ACCESS_TOKEN = "token"  
TARGET_BRANCH = "8295_master"  


# === 获取所有以 domain:: 开头的标签 ===
# 得到所有domain label
def get_domain_labels(GITLAB_URL,GROUP_PATH,ACCESS_TOKEN):
    all_mrs_info = {}
    page = 1
    per_page = 100  # 每页最大数量
    
    while True:
        # 构建API请求URL
        url = f"{GITLAB_URL}/api/v4/groups/{GROUP_PATH}/merge_requests"
         # 请求头
        headers = {
            "PRIVATE-TOKEN": ACCESS_TOKEN
        } 
        # 请求参数
        params = {
            "state": "opened",
            "scope": "all",
            "page": page,
            "per_page": per_page,
            "labels": "ready_to_review,domain::*" 
        }
        
       
        # 发送GET请求
        response = requests.get(url, params=params, headers=headers)       
        if response.status_code != 200:
            print(f"请求失败，状态码: {response.status_code}")
            print(response.text)
            break
        
        mrs = response.json()
        if not mrs:
            break
            
        # 提取所需信息
        for mr in mrs:
            mr_iid = mr.get("iid")
            project_id = mr.get("project_id")
            # 主MR的web_url
            main_url = mr.get("web_url")
            description = mr.get('description')
            labels = mr.get("labels")
            match = re.search(r'Domain:\s*([^\n,]+)', description)
            domain = match.group(1).strip() if match else None
                       
            # 只保留domain::开头的标签
            domain_labels = [label for label in mr.get('labels', []) if label.startswith('domain::')]
            
            title = mr.get("title")
            #match = re.match(r'^([A-Z]+-\d+)', jira_id)
            match = re.match(r'^(APRICOT-\d+)', title)
            if match:
                jira_id = match.group(1)
            else:
                jira_id = ""


            for label in domain_labels:
                SON_LABEL = label.replace('domain::', '')
                if SON_LABEL:
                    # TODO 得到子MR的web_url
                    son_info = get_son_url(GITLAB_URL,GROUP_PATH,ACCESS_TOKEN,SON_LABEL)
                    # 拆解出 jira_id 列表 和 son_url 列表
                    jira_ids = list(son_info.keys())
                    son_urls = list(son_info.values())
                # 以main_url为key存储数据
            all_mrs_info[main_url] = {
                "mr_iid" : mr_iid,
                "project_id" : project_id,
                "domain": domain,
                "son_urls": son_urls,  # 改为列表形式，存储所有子URL
                "jira_id": jira_ids,
                "labels": labels  # 也可以根据需要存储其他信息
            }
                  
        page += 1
        
        # 检查是否有更多页
        if 'next' not in response.links:
            break
    # 写入符合条件的 MR 到文件
    with open("filtered_mrs.json", "w", encoding="utf-8") as f:
        json.dump(all_mrs_info,f, indent=4, ensure_ascii=False)
    #print(f"总共写入 {len(all_mrs)} 个符合条件的 MR 到 filtered_mrs.json")

    return all_mrs_info


# 此函数用来查询子MR的URL，需要传入一个LABLE,
# 这个LABEL是主MR去掉domain::后面的那一截内容

def get_son_url(GITLAB_URL, GROUP_PATH, ACCESS_TOKEN, SON_LABEL):
    page = 1
    per_page = 100
    result_dict = {}

    # 编码 group_path（以防其中有斜线或空格）
    encoded_group = quote(GROUP_PATH, safe="")

    # API URL
    api_url = f"{GITLAB_URL}/api/v4/groups/{encoded_group}/merge_requests"

    headers = {
        "PRIVATE-TOKEN": ACCESS_TOKEN
    }

    # 拼接标签
    label_str = ",".join(SON_LABEL)

    while True:
        params = {
            "state": "opened",
            "labels": SON_LABEL,  # 注意传的是字符串
            "scope": "all",
            "page": page,
            "per_page": per_page
        }
        response = requests.get(api_url, headers=headers, params=params)
        if response.status_code != 200:
            print(f"请求失败: {response.status_code} - {response.text}")
            break

        mrs = response.json()
        if not mrs:
            break

        for idx, mr in enumerate(mrs):
            title = mr.get("title", "")
            son_url = mr.get("web_url", "")
            match = re.search(r'(APRICOT-\d+)', title)

            if match:
                jira_id = match.group(1)
            else:
                jira_id = f""  # 或使用 title[:20] 更友好

            result_dict[jira_id] = son_url

        # 翻页控制
        next_page = response.headers.get("X-Next-Page")
        if not next_page:
            break
        page = int(next_page)

    return result_dict


def write_mrs_to_excel(all_mrs_info, output_file="mr_review_report.xlsx"):
    """
    将all_mrs_info字典数据写入Excel表格
    
    参数:
        all_mrs_info: 包含MR信息的字典，格式为 {main_url: {domain:, son_urls:, ...}}
        output_file: 输出的Excel文件名
    """
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MR Review Report"
    
    # 定义列标题
    headers = [
        "Domain",
        "Number",
        "Multi MR or Manifest MR",
        "MRs",
        "Status",
        "Ticket",
        "Pipeline",
        "RTA Label",
        "Peer Reviews",
        "Ready_to_review Label",
        "Test Evidence",
        "MRs dependency Risk",
        "Review Result",
        "Comment"
    ]
    
    # 写入标题行并加粗
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # 写入数据
    row_num = 2
    for main_url, mr_info in all_mrs_info.items():
        # 获取MR信息
        domain = mr_info.get("domain", "")
        #main_jira_id = mr_info.get("jira_id", "")
        son_urls = mr_info.get("son_urls", [])
        jira_ids = mr_info.get("jira_id", [])
        labels = mr_info.get("labels", [])
        count = len(son_urls)
        # 判断MR类型
        mr_type = "Multi MR" if count > 0 else "Manifest MR"
        ready_to_review = "Yes" if "ready_to_review" in labels else "No"
               
        merge_rows = max(1, count)  # 至少一行
        # 合并前3列
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num + merge_rows - 1, end_column=1)
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num + merge_rows - 1, end_column=2)
        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num + merge_rows - 1, end_column=3)
        ws.cell(row=row_num, column=1, value=domain)  # Domain
        ws.cell(row=row_num, column=2, value=count)  # 子MR的数量
        ws.cell(row=row_num, column=3, value=main_url)  # 主MR

        max_len = max(len(son_urls or [""]), len(jira_ids or [""]))  # 取两个列表中较长的

        for i in range(max_len):
            current_row = row_num + i
            print(current_row)

            # 写入 son_url（如果有）
            son_url = son_urls[i] if i < len(son_urls) else ""
            ws.cell(row=current_row, column=4, value=son_url).alignment = Alignment(wrap_text=True)
            if not "ready_ng" in labels:
                ws.cell(row=current_row, column=5, value="✅")
            else:
                ws.cell(row=current_row, column=5, value="❌")

            # 写入 jira_id（如果有）
            jira_id = jira_ids[i] if i < len(jira_ids) else ""
            ws.cell(row=current_row, column=6, value=jira_id).alignment = Alignment(wrap_text=True)

            # 可选：每行都写一次 review 状态，或只写在首行（视需求）
            ws.cell(row=current_row, column=10, value=ready_to_review)

        row_num += merge_rows
        
    
    # 调整列宽
    # 设置每一列的固定宽度（比如统一为 30）
    for col in ws.columns:
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 5
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 80
        ws.column_dimensions["E"].width = 5
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 10
        ws.column_dimensions["K"].width = 10
        ws.column_dimensions["L"].width = 10
        ws.column_dimensions["M"].width = 10
        ws.column_dimensions["N"].width = 10
    
    # 保存Excel文件
    wb.save(output_file)
    print(f"报告已生成，保存为: {output_file}")



def remove_label_from_gitlab_mr(GITLAB_URL, GROUP_PATH, ACCESS_TOKEN, all_mrs_info):
    remove_label = "ready_to_review"
    for mr_data in all_mrs_info.values():
        project_id = mr_data.get("project_id")
        mr_iid = mr_data.get("mr_iid")
        url = f"{GITLAB_URL}/api/v4/projects/{project_id}/merge_requests/{mr_iid}"
    
        headers = {
            "PRIVATE-TOKEN": ACCESS_TOKEN
        }
    
        # 首先获取当前标签
        response = requests.get(url, headers=headers)
        current_labels = response.json().get("labels", [])
    
        if remove_label in current_labels:
            updated_labels = [label for label in current_labels if label != remove_label]

        # 发送更新请求
        update_resp = requests.put(url, headers=headers, json={"labels": updated_labels})
        if update_resp.status_code == 200:
            print(f"已从 MR {mr_iid} 中删除标签: {label_to_remove}")
        else:
            print(f"删除失败: {update_resp.status_code}, {update_resp.text}")



# === 主函数 ===
def main():
    try:
        # 这里会得到去掉domain::后的label
        all_mrs_info = get_domain_labels(GITLAB_URL, GROUP_PATH, ACCESS_TOKEN)
        #print(f"发现 {len(domain_labels)} 个以 'domain::' 开头的标签")

        write_mrs_to_excel(all_mrs_info, output_file="mr_review_report.xlsx")
        #remove_label_from_gitlab_mr(GITLAB_URL, GROUP_PATH, ACCESS_TOKEN, all_mrs_info)
        

    except Exception as e:
        print(f"发生错误: {e}")


if __name__ == "__main__":
    main()
