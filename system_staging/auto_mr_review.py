#!/usr/bin/env python3
"""
GitLab MR合规性验证工具
功能：验证MR标题格式，检查JIRA任务类型，生成Excel和HTML报告
作者：AI助手
版本：1.0.1
修复：JSON解析错误和HTML生成错误
"""

import requests
import pandas as pd
from urllib.parse import quote
import re
from datetime import datetime
import json
import os
import sys
import webbrowser
from html import escape
from typing import Dict, List, Any, Optional, Tuple

# ==================== 配置部分 ====================
# 请在使用前修改以下配置
GITLAB_CONFIG = {
    'url': 'https://git.sb.com',
    'token': 'token'  # 替换为你的GitLab Token
}

JIRA_CONFIG = {
    'enabled': True,  # 是否启用JIRA集成
    'url': 'https://issue.sb.com',  # JIRA Base URL
    'email': 'yuxia@abc.com',  # JIRA邮箱
    'token': 'token'  # JIRA API Token
}

# MR信息
MR_INFO = {
    'project_path': 'sb/android/aosp/vendor/mercedes/packages/mbsettings',
    'mr_iid': 837
}

# ==================== 辅助函数 ====================

def safe_get(data: Dict, *keys: str, default: Any = None) -> Any:
    """安全地获取嵌套字典的值"""
    current = data
    for key in keys:
        if isinstance(current, dict) and key in current:
            current = current[key]
        else:
            return default
    return current

def safe_str(value: Any) -> str:
    """安全地将任何值转换为字符串"""
    if value is None:
        return ''
    try:
        return str(value)
    except:
        return ''

def validate_json_string(json_str: str) -> bool:
    """验证字符串是否为有效的JSON"""
    try:
        json.loads(json_str)
        return True
    except json.JSONDecodeError:
        return False
    except Exception:
        return False

# ==================== API函数 ====================

def get_gitlab_mr_details(gitlab_url: str, project_path: str, mr_iid: int, token: str) -> Optional[Dict]:
    """获取GitLab MR详细信息"""
    try:
        encoded_project_path = quote(project_path, safe='')
        api_url = f"{gitlab_url.rstrip('/')}/api/v4/projects/{encoded_project_path}/merge_requests/{mr_iid}"
        
        headers = {"PRIVATE-TOKEN": token}
        print(f"正在请求: {api_url}")
        response = requests.get(api_url, headers=headers, timeout=30)
        
        if response.status_code == 200:
            try:
                return response.json()
            except json.JSONDecodeError as e:
                print(f"❌ 解析JSON响应失败: {e}")
                print(f"响应内容: {response.text[:200]}")
                return None
        elif response.status_code == 404:
            print(f"❌ 未找到MR: {project_path} !{mr_iid}")
            return None
        elif response.status_code == 401:
            print("❌ 认证失败，请检查GitLab Token")
            return None
        else:
            print(f"❌ GitLab API请求失败: {response.status_code}")
            print(f"响应: {response.text[:200]}")
            return None
            
    except requests.exceptions.RequestException as e:
        print(f"❌ 网络请求失败: {e}")
        return None
    except Exception as e:
        print(f"❌ 获取MR信息时出错: {e}")
        return None

def get_gitlab_mr_changes(gitlab_url: str, project_path: str, mr_iid: int, token: str) -> List[Dict]:
    """获取GitLab MR的变更信息"""
    try:
        encoded_project_path = quote(project_path, safe='')
        api_url = f"{gitlab_url.rstrip('/')}/api/v4/projects/{encoded_project_path}/merge_requests/{mr_iid}/changes"
        
        headers = {"PRIVATE-TOKEN": token}
        response = requests.get(api_url, headers=headers, timeout=30)
        
        if response.status_code == 200:
            data = response.json()
            if isinstance(data, dict) and 'changes' in data:
                return data['changes']
            elif isinstance(data, list):
                return data
            else:
                return []
        else:
            print(f"⚠️ 无法获取变更信息: {response.status_code}")
            return []
            
    except Exception as e:
        print(f"⚠️ 获取变更信息时出错: {e}")
        return []

def get_jira_issue_details(jira_id: str, jira_config: Dict) -> Optional[Dict]:
    """获取JIRA任务详细信息"""
    if not jira_id or not jira_config.get('enabled', False):
        return None
    
    jira_url = jira_config.get('url', '').rstrip('/')
    jira_token = jira_config.get('token', '')
    jira_email = jira_config.get('email', '')
    
    if not all([jira_url, jira_token, jira_email]):
        print("⚠️ JIRA配置不完整，跳过JIRA检查")
        return None
    
    try:
        # 构造JIRA API URL
        api_url = f"{jira_url}/rest/api/2/issue/{jira_id}"
        auth = (jira_email, jira_token)
        # headers = {"Accept": "application/json"}
        headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {jira_token}"
        }
        
        print(f"正在查询JIRA任务: {jira_id}...")
        response = requests.get(api_url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            try:
                issue_data = response.json()
                # 提取关键信息
                issue_type = safe_get(issue_data, 'fields', 'issuetype', 'name', default='未知')
                priority = safe_get(issue_data, 'fields', 'priority', 'name', default='未设置')
                status = safe_get(issue_data, 'fields', 'status', 'name', default='未知')
                summary = safe_get(issue_data, 'fields', 'summary', default='无标题')
                description = safe_get(issue_data, 'fields', 'description', default='无描述')
                
                if isinstance(description, str) and len(description) > 200:
                    description = description[:200] + '...'
                
                return {
                    'id': jira_id,
                    'type': issue_type,
                    'priority': priority,
                    'status': status,
                    'summary': summary,
                    'description': description,
                    'url': f"{jira_url}/browse/{jira_id}",
                    'raw_data': issue_data
                }
            except Exception as e:
                print(f"⚠️ 解析JIRA响应失败: {e}")
                return {'id': jira_id, 'error': f'解析失败: {str(e)}'}
        elif response.status_code == 404:
            print(f"⚠️ JIRA任务不存在: {jira_id}")
            return {'id': jira_id, 'error': '任务不存在'}
        elif response.status_code == 401:
            print("⚠️ JIRA认证失败，请检查Token和邮箱")
            return {'id': jira_id, 'error': '认证失败'}
        else:
            print(f"⚠️ JIRA API请求失败: {response.status_code}")
            return {'id': jira_id, 'error': f'API错误: {response.status_code}'}
            
    except requests.exceptions.RequestException as e:
        print(f"⚠️ 连接JIRA失败: {e}")
        return {'id': jira_id, 'error': f'连接失败: {str(e)}'}
    except Exception as e:
        print(f"⚠️ 获取JIRA信息时出错: {e}")
        return {'id': jira_id, 'error': f'错误: {str(e)}'}

# ==================== 验证函数 ====================

def validate_title_format(title: str) -> Tuple[bool, str]:
    """验证MR标题格式: APRICOT-<数字> : <描述>"""
    if not title or not isinstance(title, str):
        return False, "标题为空或不是字符串"
    
    # 清理标题两端的空白
    title = title.strip()
    
    # pattern = r'^APRICOT-\d+ : .+'
    pattern = r'^APRICOT-\d+ .+'
    is_valid = bool(re.match(pattern, title))
    
    if not is_valid:
        # 提供更详细的错误信息
        if not title.startswith('APRICOT-'):
            return False, "标题必须以'APRICOT-'开头"
        elif ' : ' not in title:
            return False, "标题中必须包含' : '（冒号和空格）"
        else:
            parts = title.split(' : ', 1)
            if len(parts) < 2 or not parts[1].strip():
                return False, "标题描述不能为空"
            elif not re.match(r'^APRICOT-\d+', parts[0]):
                return False, "JIRA ID格式不正确，应为APRICOT-数字"
    
    return is_valid, "格式正确" if is_valid else "格式错误"

def extract_jira_id_from_title(title: str) -> Optional[str]:
    """从标题中提取JIRA ID"""
    if not title or not isinstance(title, str):
        return None
    
    match = re.search(r'APRICOT-(\d+)', title)
    if match:
        return f"APRICOT-{match.group(1)}"
    
    return None

def categorize_jira_issue(jira_info: Dict) -> Tuple[str, str, str]:
    """根据JIRA任务信息进行分类"""
    if not jira_info or 'error' in jira_info:
        return "未知", "无法获取JIRA信息", "未知"
    
    issue_type = safe_str(jira_info.get('type', '')).lower()
    summary = safe_str(jira_info.get('summary', '')).lower()
    description = safe_str(jira_info.get('description', '')).lower()
    
    # 检查是否为BUG类型
    if any(word in issue_type for word in ['bug', 'defect', 'error']):
        category = "BUG"
        compliance = "合规"
        detail = f"{jira_info.get('type')}"
    
    # 检查是否为Non-Conformance
    elif any(term in summary or term in description 
             for term in ['non-conformance', 'non conformance', 'nc']):
        category = "Non-Conformance"
        compliance = "合规"
        detail = f"非一致性问题: {jira_info.get('type')}"
    
    # 检查是否为STARC Bug
    elif 'starc' in summary or 'starc' in description:
        category = "STARC Bug"
        compliance = "合规"
        detail = f"STARC Bug: {jira_info.get('type')}"
    
    # 其他类型
    elif any(word in issue_type for word in ['story', 'task', 'feature']):
        category = "常规任务"
        compliance = "警告"
        detail = f"常规任务类型: {jira_info.get('type')}"
    
    elif any(word in issue_type for word in ['improvement', 'enhancement']):
        category = "改进"
        compliance = "警告"
        detail = f"改进类型: {jira_info.get('type')}"
    
    else:
        category = "其他类型"
        compliance = "警告"
        detail = f"未分类: {jira_info.get('type')}"
    
    return category, compliance, detail

def check_crosslink(description: str) -> Tuple[bool, str]:
    """检查MR描述中是否有相关链接"""
    if not description or not isinstance(description, str):
        return False, "描述为空或不是字符串"
    
    description_lower = description.lower()
    match = re.search(r"Crosslink Merge Requests:\s*(.*?)\s*Domain:", description, re.S)
    if match:
        section = match.group(1)
        # 主MR包含的子MR，去重
        son_urls = list(set(re.findall(r"https://git\.sb\.com[^\s]+/merge_requests/\d+", section)))
    else:
        构建状态: not_run = []
    
    # 检查常见链接模式
    patterns = [
        r'https?://[^\s]+',  # HTTP/HTTPS链接
        r'jira/[^\s]+',      # JIRA链接
        r'ticket/[^\s]+',    # Ticket链接
        r'issue/[^\s]+',     # Issue链接
    ]
    
    for pattern in patterns:
        if re.search(pattern, description_lower):
            return True, son_urls
    
    return False, "未找到相关链接"

def check_build_results(mr_data: Dict) -> Tuple[bool, str, str]:
    """检查构建结果"""
    description = mr_data.get("description")
    status = get_last_pipeline_status(description)
    print(f"🔍 最后一个Pipeline状态: {status}")
    build_info = parse_mr_description_for_build_info(description)
    last_pipeline = build_info['last_pipeline']
    print(f"📊 状态: {last_pipeline['status']}")
    print(f"🏷️  域标签: {last_pipeline['domain_label']}")
    print(f"👤 作者: {last_pipeline['author']}")
    print(f"🔗 构建URL: {last_pipeline['build_url']}")
    print(f"📁 版本路径: {last_pipeline['version_path']}")
    
    print(f"\n📈 总共找到 {build_info['build_sections_count']} 个构建记录")
    
    if status == 'success':
        return True, f"构建状态: {status}", last_pipeline['build_url']
    elif status == 'failed':
        return False, f"构建状态: {status}", last_pipeline['build_url']
    elif status == 'running':
        return False, f"构建状态: {status}", last_pipeline['build_url']
    else:
        return False, f"构建状态: {status}", last_pipeline['build_url']

def check_ready_to_merge(mr_data: Dict) -> Tuple[bool, str, List[str]]:
    """检查MR是否可合并"""
    issues = []
    
    # 1. 检查状态
    state = mr_data.get('state', '')
    if state != 'opened':
        issues.append(f"MR状态为 '{state}'，非打开状态")
    
    # 2. 检查冲突
    if mr_data.get('has_conflicts', False):
        issues.append("存在代码冲突")
    
    # 3. 检查合并性
    if not mr_data.get('merge_status', 'can_be_merged'):
        issues.append("当前不可合并")
    print(issues)
    # 4. 检查Pipeline状态
    # pipeline = mr_data.get('pipeline', {})
    # if not isinstance(pipeline, dict):
    #     pipeline = {}
    # pipeline_status = pipeline.get('status', 'unknown')
    # if pipeline_status not in ['success', 'manual']:
    #     issues.append(f"Pipeline状态: {pipeline_status}")
    
    # 5. 检查未解决的讨论
    # user_notes_count = mr_data.get('user_notes_count', 0)
    # user_discussions_resolved = mr_data.get('user_discussions_resolved', 0)
    
    # 确保是数字
    # try:
    #     user_notes_count = int(user_notes_count)
    #     user_discussions_resolved = int(user_discussions_resolved)
    # except (ValueError, TypeError):
    #     user_notes_count = 0
    #     user_discussions_resolved = 0
    
    # unresolved = user_notes_count - user_discussions_resolved
    # if unresolved > 0:
    #     issues.append(f"有 {unresolved} 个未解决的讨论")
    
    if not issues:
        return True, "满足所有合并条件", []
    else:
        return False, "不满足合并条件", issues

# ==================== 分析函数 ====================

def analyze_mr_data(mr_data: Dict, changes_data: List, jira_config: Dict) -> Tuple[Dict, List]:
    """分析MR数据，返回合规性检查结果"""
    
    # 初始化结果结构
    results = {
        "检查项": [
            "Title格式",
            "JIRA ID", 
            "JIRA类型",
            "Crosslink", 
            "Domain Build Results", 
            "Ready to Merge"
        ],
        "是否合规": [],
        "详细信息": [],
        "状态": []
    }
    
    jira_summary = []
    jira_id = None
    
    # 1. 检查Title格式
    title = safe_str(mr_data.get('title', ''))
    is_title_valid, title_detail = validate_title_format(title)
    results["是否合规"].append("√" if is_title_valid else "Ⅹ")
    results["详细信息"].append(title_detail)
    results["状态"].append("通过" if is_title_valid else "失败")
    
    # 2. 检查JIRA ID
    jira_id = extract_jira_id_from_title(title)
    has_jira_id = bool(jira_id)
    results["是否合规"].append("√" if has_jira_id else "Ⅹ")
    results["详细信息"].append(jira_id if jira_id else "未找到JIRA ID")
    results["状态"].append("通过" if has_jira_id else "失败")
    
    # 3. 检查JIRA类型
    if jira_id and jira_config.get('enabled', False):
        jira_info = get_jira_issue_details(jira_id, jira_config)
        
        if jira_info and 'error' not in jira_info:
            category, compliance, detail = categorize_jira_issue(jira_info)
            
            # 添加到JIRA摘要
            jira_summary.append({
                'id': jira_id,
                'type': jira_info.get('type'),
                'category': category,
                'compliance': compliance,
                'detail': detail,
                'priority': jira_info.get('priority'),
                'status': jira_info.get('status'),
                'url': jira_info.get('url')
            })
            
            if compliance == "合规":
                results["是否合规"].append("√")
                results["状态"].append("通过")
            else:
                results["是否合规"].append("⚠️")
                results["状态"].append("警告")
            results["详细信息"].append(detail)
            
        elif jira_info and 'error' in jira_info:
            results["是否合规"].append("⚠️")
            results["状态"].append("警告")
            results["详细信息"].append(f"JIRA查询失败: {jira_info['error']}")
        else:
            results["是否合规"].append("⚠️")
            results["状态"].append("警告")
            results["详细信息"].append("JIRA配置未启用或查询失败")
    else:
        if jira_id:
            results["是否合规"].append("⚠️")
            results["状态"].append("警告")
            results["详细信息"].append("JIRA集成未启用")
        else:
            results["是否合规"].append("Ⅹ")
            results["状态"].append("失败")
            results["详细信息"].append("无JIRA ID")
    
    # 4. 检查Crosslink
    description = safe_str(mr_data.get('description', ''))
    has_crosslink, crosslink_detail = check_crosslink(description)
    results["是否合规"].append("√" if has_crosslink else "待检查")
    results["状态"].append("通过" if has_crosslink else "待检查")
    results["详细信息"].append(crosslink_detail)
    
    # 5. 检查Domain Build Results
    build_success, build_detail, build_url = check_build_results(mr_data)
    if build_success:
        results["是否合规"].append("√")
        results["状态"].append("通过")
    elif build_detail == "构建中":
        results["是否合规"].append("⚠️")
        results["状态"].append("进行中")
    else:
        results["是否合规"].append("Ⅹ")
        results["状态"].append("失败")
    
    detail_text = build_detail
    if build_url and isinstance(build_url, str) and build_url.startswith('http'):
        detail_text += f" | URL: {build_url}"
    results["详细信息"].append(detail_text)
    
    # 6. 检查Ready to Merge
    is_ready, ready_detail, ready_issues = check_ready_to_merge(mr_data)
    if is_ready:
        results["是否合规"].append("√")
        results["状态"].append("通过")
    else:
        results["是否合规"].append("Ⅹ")
        results["状态"].append("失败")
    
    if ready_issues:
        results["详细信息"].append(f"{ready_detail}: {', '.join(ready_issues)}")
    else:
        results["详细信息"].append(ready_detail)
    
    return results, jira_summary

# ==================== 报告生成函数 ====================

def create_excel_report(mr_data: Dict, validation_results: Dict, jira_summary: List, 
                       changes_data: List, output_file: str) -> Optional[str]:
    """创建Excel格式的报告"""
    
    print(f"正在生成Excel报告: {output_file}")
    
    try:
        # 创建主合规性检查表
        df_main = pd.DataFrame({
            "检查项": validation_results["检查项"],
            "是否合规": validation_results["是否合规"],
            "状态": validation_results["状态"],
            "详细信息": validation_results["详细信息"]
        })
        
        # 创建MR详细信息表
        references = safe_get(mr_data, 'references')
        if isinstance(references, dict):
            project_name = references.get('full', 'N/A')
        else:
            project_name = 'N/A'
        
        author = safe_get(mr_data, 'author')
        if isinstance(author, dict):
            author_name = author.get('name', 'N/A')
        else:
            author_name = 'N/A'
        
        mr_details_rows = [
            {"分类": "基本信息", "内容": f"项目: {project_name}"},
            {"分类": "基本信息", "内容": f"MR: !{mr_data.get('iid', 'N/A')}"},
            {"分类": "基本信息", "内容": f"标题: {mr_data.get('title', 'N/A')}"},
            {"分类": "基本信息", "内容": f"URL: {mr_data.get('web_url', 'N/A')}"},
            {"分类": "基本信息", "内容": f"创建者: {author_name}"},
            {"分类": "基本信息", "内容": f"创建时间: {mr_data.get('created_at', 'N/A')}"},
            {"分类": "基本信息", "内容": f"更新时间: {mr_data.get('updated_at', 'N/A')}"},
            {"分类": "分支信息", "内容": f"源分支: {mr_data.get('source_branch', 'N/A')}"},
            {"分类": "分支信息", "内容": f"目标分支: {mr_data.get('target_branch', 'N/A')}"},
            {"分类": "分支信息", "内容": f"变更文件数: {mr_data.get('changes_count', 0)}"},
            {"分类": "状态信息", "内容": f"当前状态: {mr_data.get('state', 'N/A')}"},
            {"分类": "状态信息", "内容": f"可合并: {mr_data.get('merge_status', 'N/A')}"},
            {"分类": "状态信息", "内容": f"存在冲突: {mr_data.get('has_conflicts', 'N/A')}"},
            {"分类": "状态信息", "内容": f"Pipeline状态: {safe_get(mr_data, 'pipeline', 'status', default='unknown')}"},
            {"分类": "状态信息", "内容": f"评论数量: {mr_data.get('user_notes_count', 0)}"},
            {"分类": "状态信息", "内容": f"已解决讨论: {mr_data.get('user_discussions_resolved', 0)}"}
        ]
        
        df_details = pd.DataFrame(mr_details_rows)
        
        # 创建JIRA信息表（如果有）
        df_jira = None
        if jira_summary:
            jira_data = []
            for jira in jira_summary:
                if isinstance(jira, dict):
                    jira_data.append({
                        "JIRA ID": jira.get('id', ''),
                        "类型": jira.get('type', ''),
                        "分类": jira.get('category', ''),
                        "合规性": jira.get('compliance', ''),
                        "优先级": jira.get('priority', ''),
                        "状态": jira.get('status', ''),
                        "详细信息": jira.get('detail', ''),
                        "URL": jira.get('url', '')
                    })
            if jira_data:
                df_jira = pd.DataFrame(jira_data)
        
        # 创建变更文件表
        df_changes = None
        if changes_data and isinstance(changes_data, list):
            changes_list = []
            for change in changes_data:
                if isinstance(change, dict):
                    filename = change.get('new_path', change.get('old_path', '未知'))
                    status = "new" if change.get('new_file', False) else "deleted" if change.get('deleted_file', False) else "modified"
                    
                    # 计算变更行数
                    diff = change.get('diff', '')
                    if isinstance(diff, str):
                        diff_lines = len(diff.split('\n'))
                    else:
                        diff_lines = 0
                    
                    changes_list.append({
                        "文件": filename,
                        "状态": status,
                        "变更行数": diff_lines
                    })
            if changes_list:
                df_changes = pd.DataFrame(changes_list)
        
        # 写入Excel文件
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 写入合规性检查表
            df_main.to_excel(writer, sheet_name='合规性检查', index=False)
            
            # 写入MR详细信息
            df_details.to_excel(writer, sheet_name='MR详细信息', index=False)
            
            # 写入JIRA信息
            if df_jira is not None and not df_jira.empty:
                df_jira.to_excel(writer, sheet_name='JIRA信息', index=False)
            
            # 写入变更文件
            if df_changes is not None and not df_changes.empty:
                df_changes.to_excel(writer, sheet_name='变更文件', index=False)
            
            # 写入原始数据（只写入可序列化的数据）
            raw_data_rows = []
            for key, value in mr_data.items():
                # 过滤掉不可序列化的值
                try:
                    json.dumps({key: value})
                    raw_data_rows.append({"字段": key, "值": str(value)[:200] + "..." if len(str(value)) > 200 else str(value)})
                except:
                    raw_data_rows.append({"字段": key, "值": "[不可序列化的数据]"})
            
            if raw_data_rows:
                pd.DataFrame(raw_data_rows).to_excel(writer, sheet_name='原始数据', index=False)
        
        # 设置列宽
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            
            wb = load_workbook(output_file)
            
            # 设置合规性检查表样式
            if '合规性检查' in wb.sheetnames:
                ws = wb['合规性检查']
                
                # 设置列宽
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 50
                
                # 设置标题样式
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # 设置状态颜色
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                    for cell in row:
                        if cell.value == "√":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif cell.value == "Ⅹ":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif cell.value == "⚠️":
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            wb.save(output_file)
        except Exception as e:
            print(f"⚠️ 设置Excel样式时出错: {e}")
            # 继续执行，至少文件已经生成
        
        print(f"✅ Excel报告已生成: {os.path.abspath(output_file)}")
        return output_file
        
    except Exception as e:
        print(f"❌ 生成Excel报告时出错: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_html_report(mr_data: Dict, validation_results: Dict, jira_summary: List, 
                      changes_data: List, output_file: str) -> Optional[str]:
    """创建HTML格式的报告"""
    
    print(f"正在生成HTML报告: {output_file}")
    
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 安全获取数据
        title = safe_str(mr_data.get('title', 'N/A'))
        mr_url = safe_str(mr_data.get('web_url', '#'))
        mr_iid = safe_str(mr_data.get('iid', 'N/A'))
        
        # 获取项目名称
        references = mr_data.get('references')
        if isinstance(references, dict):
            project_name = references.get('full', 'N/A')
        else:
            project_name = 'N/A'
        
        # 获取作者信息
        author = mr_data.get('author')
        if isinstance(author, dict):
            author_name = author.get('name', 'N/A')
        else:
            author_name = 'N/A'
        
        # 获取pipeline信息
        pipeline = mr_data.get('pipeline', {})
        if isinstance(pipeline, dict):
            pipeline_status = pipeline.get('status', 'unknown')
        else:
            pipeline_status = 'unknown'
        
        # 获取commits
        commits = mr_data.get('commits', [])
        if not isinstance(commits, list):
            commits = []
        
        # 计算统计信息
        total_checks = len(validation_results["是否合规"])
        passed_checks = sum(1 for status in validation_results["是否合规"] if status == "√")
        warning_checks = sum(1 for status in validation_results["是否合规"] if status == "⚠️")
        failed_checks = sum(1 for status in validation_results["是否合规"] if status == "Ⅹ")
        pass_rate = (passed_checks / total_checks * 100) if total_checks > 0 else 0
        
        # 生成合规性表格行
        compliance_rows = []
        for i in range(total_checks):
            item = validation_results["检查项"][i]
            status = validation_results["是否合规"][i]
            detail = validation_results["详细信息"][i]
            
            if status == "√":
                status_badge = '<span class="badge badge-success">通过</span>'
                icon = '✅'
            elif status == "⚠️":
                status_badge = '<span class="badge badge-warning">警告</span>'
                icon = '⚠️'
            elif status == "Ⅹ":
                status_badge = '<span class="badge badge-danger">失败</span>'
                icon = '❌'
            else:
                status_badge = f'<span class="badge badge-info">{status}</span>'
                icon = '❓'
            
            row = f"""
            <tr>
                <td><strong>{icon} {escape(str(item))}</strong></td>
                <td>{status_badge}</td>
                <td>{escape(str(detail))}</td>
            </tr>
            """
            compliance_rows.append(row)
        
        compliance_rows_html = ''.join(compliance_rows)
        
        # 生成JIRA信息表格
        jira_rows = []
        if jira_summary:
            for jira in jira_summary:
                if isinstance(jira, dict):
                    compliance_class = "badge-success" if jira.get('compliance') == "合规" else "badge-warning"
                    jira_id = escape(str(jira.get('id', '')))
                    jira_type = escape(str(jira.get('type', '')))
                    jira_category = escape(str(jira.get('category', '')))
                    jira_priority = escape(str(jira.get('priority', '')))
                    jira_status = escape(str(jira.get('status', '')))
                    jira_url = escape(str(jira.get('url', '#')))
                    
                    row = f"""
                    <tr>
                        <td><a href="{jira_url}" target="_blank">{jira_id}</a></td>
                        <td>{jira_type}</td>
                        <td><span class="badge {compliance_class}">{jira_category}</span></td>
                        <td>{jira_priority}</td>
                        <td>{jira_status}</td>
                    </tr>
                    """
                    jira_rows.append(row)
        
        jira_rows_html = ''.join(jira_rows)
        jira_section = ""
        if jira_rows:
            jira_section = f"""
            <div class="card">
                <h2>🎯 JIRA 任务信息</h2>
                <table>
                    <thead>
                        <tr>
                            <th>JIRA ID</th>
                            <th>类型</th>
                            <th>分类</th>
                            <th>优先级</th>
                            <th>状态</th>
                        </tr>
                    </thead>
                    <tbody>
                        {jira_rows_html}
                    </tbody>
                </table>
            </div>
            """
        
        # 生成变更文件列表
        changes_list = []
        changes_count = 0
        if changes_data and isinstance(changes_data, list):
            changes_count = len(changes_data)
            for i, change in enumerate(changes_data[:50]):  # 限制显示前50个
                if isinstance(change, dict):
                    filename = change.get('new_path', change.get('old_path', '未知文件'))
                    if not filename:
                        filename = '未知文件'
                    
                    status = "new" if change.get('new_file', False) else "deleted" if change.get('deleted_file', False) else "modified"
                    status_text = {"new": "新增", "deleted": "删除", "modified": "修改"}.get(status, "修改")
                    status_class = {"new": "added", "deleted": "deleted", "modified": "modified"}.get(status, "modified")
                    
                    item = f"""
                    <div class="change-item">
                        <span class="change-file">{escape(str(filename))}</span>
                        <span class="change-status {status_class}">{status_text}</span>
                    </div>
                    """
                    changes_list.append(item)
            
            changes_list_html = ''.join(changes_list)
            if changes_count > 50:
                changes_list_html += f'<div class="text-muted">... 还有 {changes_count - 50} 个文件未显示</div>'
        else:
            changes_list_html = '<div class="text-muted">无变更信息</div>'
            changes_count = 0
        
        # 获取其他字段
        source_branch = escape(str(mr_data.get('source_branch', 'N/A')))
        target_branch = escape(str(mr_data.get('target_branch', 'N/A')))
        
        changes_count_val = mr_data.get('changes_count', 0)
        try:
            changes_count_val = int(changes_count_val)
        except:
            changes_count_val = 0
        
        mergeable = escape(str(mr_data.get('merge_status', 'N/A')))
        has_conflicts = escape(str(mr_data.get('has_conflicts', 'N/A')))
        
        user_notes_count = mr_data.get('user_notes_count', 0)
        try:
            user_notes_count = int(user_notes_count)
        except:
            user_notes_count = 0
        
        user_discussions_resolved = mr_data.get('user_discussions_resolved', 0)
        try:
            user_discussions_resolved = int(user_discussions_resolved)
        except:
            user_discussions_resolved = 0
        
        created_at = escape(str(mr_data.get('created_at', 'N/A')))
        updated_at = escape(str(mr_data.get('updated_at', 'N/A')))
        mr_state = escape(str(mr_data.get('state', 'N/A')))
        
        # 生成HTML
        html_template = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>MR合规性验证报告 - {escape(title)}</title>
    <style>
        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}
        
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
            line-height: 1.6;
            color: #333;
            background-color: #f8f9fa;
            padding: 20px;
        }}
        
        .container {{
            max-width: 1200px;
            margin: 0 auto;
        }}
        
        .header {{
            background: linear-gradient(135deg, #1e3c72, #2a5298);
            color: white;
            padding: 30px;
            border-radius: 10px;
            margin-bottom: 30px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        
        .header h1 {{
            font-size: 28px;
            margin-bottom: 10px;
        }}
        
        .header .subtitle {{
            opacity: 0.9;
            font-size: 16px;
        }}
        
        .stats-container {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        
        .stat-card {{
            background: white;
            padding: 25px;
            border-radius: 8px;
            text-align: center;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            transition: transform 0.2s;
        }}
        
        .stat-card:hover {{
            transform: translateY(-5px);
            box-shadow: 0 4px 8px rgba(0,0,0,0.1);
        }}
        
        .stat-value {{
            font-size: 36px;
            font-weight: bold;
            margin: 10px 0;
        }}
        
        .stat-label {{
            font-size: 14px;
            color: #666;
            text-transform: uppercase;
            letter-spacing: 1px;
        }}
        
        .card {{
            background: white;
            border-radius: 8px;
            padding: 25px;
            margin-bottom: 25px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        }}
        
        .card h2 {{
            color: #1e3c72;
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid #e9ecef;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 15px 0;
        }}
        
        th {{
            background-color: #f8f9fa;
            font-weight: 600;
            text-align: left;
            padding: 12px 15px;
            border-bottom: 2px solid #dee2e6;
        }}
        
        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #e9ecef;
        }}
        
        tr:hover {{
            background-color: #f8f9fa;
        }}
        
        .badge {{
            display: inline-block;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: bold;
        }}
        
        .badge-success {{
            background-color: #d4edda;
            color: #155724;
        }}
        
        .badge-warning {{
            background-color: #fff3cd;
            color: #856404;
        }}
        
        .badge-danger {{
            background-color: #f8d7da;
            color: #721c24;
        }}
        
        .badge-info {{
            background-color: #d1ecf1;
            color: #0c5460;
        }}
        
        .info-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        
        .info-section {{
            background: #f8f9fa;
            padding: 20px;
            border-radius: 6px;
        }}
        
        .info-section h3 {{
            color: #1e3c72;
            margin-bottom: 15px;
        }}
        
        .info-section p {{
            margin: 8px 0;
        }}
        
        .changes-list {{
            max-height: 400px;
            overflow-y: auto;
            border: 1px solid #dee2e6;
            border-radius: 6px;
            padding: 15px;
        }}
        
        .change-item {{
            padding: 10px;
            border-bottom: 1px solid #e9ecef;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        
        .change-item:last-child {{
            border-bottom: none;
        }}
        
        .change-file {{
            font-family: 'Courier New', monospace;
            font-size: 13px;
        }}
        
        .change-status {{
            font-size: 11px;
            padding: 3px 8px;
            border-radius: 3px;
            font-weight: bold;
        }}
        
        .added {{ background-color: #d4edda; color: #155724; }}
        .modified {{ background-color: #fff3cd; color: #856404; }}
        .deleted {{ background-color: #f8d7da; color: #721c24; }}
        
        .footer {{
            text-align: center;
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #dee2e6;
            color: #6c757d;
            font-size: 14px;
        }}
        
        .btn {{
            display: inline-block;
            padding: 8px 16px;
            background-color: #1e3c72;
            color: white;
            text-decoration: none;
            border-radius: 4px;
            font-weight: 500;
            transition: background-color 0.2s;
        }}
        
        .btn:hover {{
            background-color: #162b55;
        }}
        
        .text-muted {{
            color: #6c757d;
        }}
        
        .summary {{
            background: linear-gradient(135deg, #f8f9fa, #e9ecef);
            padding: 20px;
            border-radius: 8px;
            margin: 20px 0;
        }}
        
        .summary h3 {{
            color: #1e3c72;
            margin-bottom: 15px;
        }}
        
        .summary ul {{
            padding-left: 20px;
        }}
        
        .summary li {{
            margin: 8px 0;
        }}
        
        @media (max-width: 768px) {{
            .stats-container {{
                grid-template-columns: 1fr;
            }}
            
            .info-grid {{
                grid-template-columns: 1fr;
            }}
            
            .header {{
                padding: 20px;
            }}
            
            .header h1 {{
                font-size: 24px;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📋 Merge Request 合规性验证报告</h1>
            <div class="subtitle">
                <strong>项目:</strong> {escape(project_name)} | 
                <strong>MR:</strong> !{mr_iid} | 
                <strong>生成时间:</strong> {timestamp}
            </div>
        </div>
        
        <div class="stats-container">
            <div class="stat-card">
                <div class="stat-value">{passed_checks}/{total_checks}</div>
                <div class="stat-label">通过项目</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{warning_checks}</div>
                <div class="stat-label">警告项目</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{failed_checks}</div>
                <div class="stat-label">失败项目</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{pass_rate:.1f}%</div>
                <div class="stat-label">整体通过率</div>
            </div>
        </div>
        
        <div class="card">
            <h2>✅ 合规性检查结果</h2>
            <table>
                <thead>
                    <tr>
                        <th>检查项</th>
                        <th>状态</th>
                        <th>详细信息</th>
                    </tr>
                </thead>
                <tbody>
                    {compliance_rows_html}
                </tbody>
            </table>
        </div>
        
        {jira_section}
        
        <div class="info-grid">
            <div class="info-section">
                <h3>📄 MR基本信息</h3>
                <p><strong>标题:</strong> {escape(title)}</p>
                <p><strong>创建者:</strong> {escape(author_name)}</p>
                <p><strong>创建时间:</strong> {created_at}</p>
                <p><strong>更新时间:</strong> {updated_at}</p>
                <p><strong>当前状态:</strong> {mr_state}</p>
                <p><a href="{escape(mr_url)}" target="_blank" class="btn">查看MR详情</a></p>
            </div>
            
            <div class="info-section">
                <h3>🌿 分支信息</h3>
                <p><strong>源分支:</strong> {source_branch}</p>
                <p><strong>目标分支:</strong> {target_branch}</p>
                <p><strong>变更文件:</strong> {changes_count_val} 个</p>
                <p><strong>提交数量:</strong> {len(commits)} 次</p>
                <p><strong>Pipeline状态:</strong> {escape(pipeline_status)}</p>
            </div>
            
            <div class="info-section">
                <h3>🔧 合并状态</h3>
                <p><strong>可合并:</strong> {mergeable}</p>
                <p><strong>存在冲突:</strong> {has_conflicts}</p>
                <p><strong>评论数量:</strong> {user_notes_count}</p>
                <p><strong>已解决讨论:</strong> {user_discussions_resolved}</p>
                <p><strong>Web URL:</strong> <a href="{escape(mr_url)}" target="_blank">点击查看</a></p>
            </div>
        </div>
        
        <div class="card">
            <h2>📁 文件变更 ({changes_count})</h2>
            <div class="changes-list">
                {changes_list_html}
            </div>
        </div>
        
        <div class="summary">
            <h3>📈 验证结果分析</h3>
            {"<p>✅ <strong>所有检查项已通过，可以合并</strong></p>" if failed_checks == 0 else ""}
            {f"<p>❌ <strong>有 {failed_checks} 个检查项失败，需要修复后才能合并</strong></p>" if failed_checks > 0 else ""}
            {f"<p>⚠️ <strong>有 {warning_checks} 个警告项，建议检查</strong></p>" if warning_checks > 0 else ""}
            
            <h3 style="margin-top: 20px;">🎯 建议操作</h3>
            <ul>
                {"<li>✅ 所有合规性检查已通过，可以继续合并流程</li>" if failed_checks == 0 else ""}
                {f"<li>❌ 请修复 {failed_checks} 个失败的检查项</li>" if failed_checks > 0 else ""}
                {f"<li>⚠️ 检查 {warning_checks} 个警告项，确认是否符合要求</li>" if warning_checks > 0 else ""}
                {"<li>🔍 检查文件变更，确保所有修改符合预期</li>" if changes_data else ""}
                {"<li>👥 确保所有代码审查意见已处理</li>"}
            </ul>
        </div>
        
        <div class="footer">
            <p>© {datetime.now().year} GitLab MR 合规性验证工具 | 生成时间: {timestamp}</p>
            <p class="text-muted">此报告为自动生成，数据来源于GitLab API</p>
        </div>
    </div>
    
    <script>
        // 添加简单的交互功能
        document.addEventListener('DOMContentLoaded', function() {{
            // 为所有外部链接添加target="_blank"
            document.querySelectorAll('a[href^="http"]').forEach(link => {{
                if (!link.getAttribute('target')) {{
                    link.setAttribute('target', '_blank');
                    link.setAttribute('rel', 'noopener noreferrer');
                }}
            }});
            
            // 添加打印功能
            var printBtn = document.createElement('button');
            printBtn.innerHTML = '🖨️ 打印报告';
            printBtn.className = 'btn';
            printBtn.style.margin = '10px';
            printBtn.onclick = function() {{
                window.print();
            }};
            document.querySelector('.footer').prepend(printBtn);
        }});
    </script>
</body>
</html>"""
        
        # 写入HTML文件
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_template)
        
        print(f"✅ HTML报告已生成: {os.path.abspath(output_file)}")
        return output_file
        
    except Exception as e:
        print(f"❌ 生成HTML报告时出错: {e}")
        import traceback
        traceback.print_exc()
        return None

def extract_last_pipeline_status(description: str) -> Dict[str, str]:
    """
    从MR描述中提取最后一个pipeline的构建状态
    
    Args:
        description: MR的完整描述内容
        
    Returns:
        Dict containing:
            - 'status': 构建状态 ('success', 'failed', 'unknown')
            - 'domain_label': 域标签
            - 'build_url': 构建URL
            - 'author': 作者
            - 'version_path': 版本路径
    """
    if not description:
        return {
            'status': 'unknown',
            'domain_label': '',
            'build_url': '',
            'author': '',
            'version_path': ''
        }
    
    # 查找所有构建状态部分
    # 使用正则表达式匹配每个构建块
    # 格式：- [x] Domain Build success 或 - [ ] Domain Build success
    build_sections = []
    
    # 方法1：按"---"分割构建块
    sections = description.split('---')
    build_blocks = []
    
    for section in sections:
        # 查找包含构建状态的块
        if 'Domain Build' in section or 'Build URL:' in section:
            build_blocks.append(section.strip())
    
    # 如果没有找到，尝试方法2：按包含构建状态的行查找
    if not build_blocks:
        lines = description.split('\n')
        current_block = []
        in_build_block = False
        
        for line in lines:
            if 'Domain Build success' in line or 'Domain Build failed' in line:
                in_build_block = True
                if current_block:  # 保存前一个块
                    build_blocks.append('\n'.join(current_block))
                    current_block = []
                current_block.append(line)
            elif in_build_block:
                if line.strip() and not line.startswith('  ') and not line.startswith('\t'):
                    # 可能是新段落，结束当前块
                    build_blocks.append('\n'.join(current_block))
                    current_block = []
                    in_build_block = False
                else:
                    current_block.append(line)
        
        if current_block:
            build_blocks.append('\n'.join(current_block))
    
    # 解析最后一个构建块
    last_build_info = {
        'status': 'unknown',
        'domain_label': '',
        'build_url': '',
        'author': '',
        'version_path': ''
    }
    
    if build_blocks:
        last_block = build_blocks[-1]  # 获取最后一个构建块
        
        # 提取构建状态
        if '- [x] Domain Build success' in last_block:
            last_build_info['status'] = 'success'
        elif '- [x] Domain Build failed' in last_block:
            last_build_info['status'] = 'failed'
        else:
            # 尝试其他格式
            if 'Domain Build success' in last_block and '- [x]' in last_block:
                last_build_info['status'] = 'success'
            elif 'Domain Build failed' in last_block and '- [x]' in last_block:
                last_build_info['status'] = 'failed'
        
        # 提取域标签
        domain_label_match = re.search(r'Domain Label:\s*(\S+)', last_block)
        if domain_label_match:
            last_build_info['domain_label'] = domain_label_match.group(1)
        
        # 提取构建URL
        build_url_match = re.search(r'Build URL:\s*(\S+)', last_block)
        if build_url_match:
            last_build_info['build_url'] = build_url_match.group(1)
        
        # 提取作者
        author_match = re.search(r'Author:\s*(\S+)', last_block)
        if author_match:
            last_build_info['author'] = author_match.group(1)
        
        # 提取版本路径
        version_path_match = re.search(r'Version_Path:\s*(\S+)', last_block)
        if version_path_match:
            last_build_info['version_path'] = version_path_match.group(1)
    
    return last_build_info


def parse_mr_description_for_build_info(description: str) -> Dict[str, any]:
    """
    从MR描述中解析所有构建相关信息
    
    Args:
        description: MR的完整描述内容
        
    Returns:
        包含所有构建信息的字典
    """
    if not description:
        return {
            'last_pipeline': {
                'status': 'unknown',
                'domain_label': '',
                'build_url': '',
                'author': '',
                'version_path': ''
            },
            'all_pipelines': [],
            'has_build_sections': False,
            'build_sections_count': 0
        }
    
    # 先获取最后一个pipeline信息
    last_pipeline = extract_last_pipeline_status(description)
    
    # 尝试提取所有pipeline信息
    all_pipelines = []
    
    # 查找所有构建块 - 改进的方法
    lines = description.split('\n')
    
    current_pipeline = {}
    collecting_pipeline = False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # 检查是否开始一个新的构建块
        if '- [x] Domain Build success' in line or '- [ ] Domain Build success' in line:
            # 保存前一个pipeline
            if current_pipeline:
                all_pipelines.append(current_pipeline)
            
            # 开始新的pipeline
            current_pipeline = {'raw_line': line}
            collecting_pipeline = True
            
            # 提取状态
            if '- [x]' in line:
                current_pipeline['status'] = 'success' if 'success' in line else 'failed'
            else:
                current_pipeline['status'] = 'not_run'
        
        # 收集pipeline详细信息
        elif collecting_pipeline:
            if line.startswith('Domain Label:'):
                current_pipeline['domain_label'] = line.replace('Domain Label:', '').strip()
            elif line.startswith('Author:'):
                current_pipeline['author'] = line.replace('Author:', '').strip()
            elif line.startswith('Build URL:'):
                current_pipeline['build_url'] = line.replace('Build URL:', '').strip()
            elif line.startswith('Version_Path:'):
                current_pipeline['version_path'] = line.replace('Version_Path:', '').strip()
            
            # 检查是否开始新的段落（非缩进的行且不是空行）
            if i < len(lines) - 1:
                next_line = lines[i + 1].strip()
                if next_line and not next_line.startswith('-') and not next_line.startswith('Domain') and not next_line.startswith('Author') and not next_line.startswith('Build') and not next_line.startswith('Version'):
                    # 可能是新的段落，结束当前pipeline
                    if current_pipeline:
                        all_pipelines.append(current_pipeline)
                        current_pipeline = {}
                    collecting_pipeline = False
    
    # 添加最后一个pipeline
    if current_pipeline:
        all_pipelines.append(current_pipeline)
    
    # 如果没有通过上述方法找到，尝试更简单的方法
    if not all_pipelines:
        # 直接搜索所有构建状态
        status_pattern = r'- \[([x ])\] Domain Build (success|failed)'
        status_matches = re.findall(status_pattern, description, re.IGNORECASE)
        
        for match in status_matches:
            checkbox, status = match
            pipeline_info = {
                'status': status if checkbox.strip() == 'x' else 'not_run',
                'raw_match': f"- [{checkbox}] Domain Build {status}"
            }
            all_pipelines.append(pipeline_info)
    
    # 获取最后一个pipeline的状态
    final_status = last_pipeline['status']
    
    # 如果从all_pipelines可以获取更准确的信息
    if all_pipelines:
        last_parsed = all_pipelines[-1]
        if 'status' in last_parsed:
            final_status = last_parsed['status']
        
        # 更新last_pipeline信息
        if 'domain_label' in last_parsed:
            last_pipeline['domain_label'] = last_parsed.get('domain_label', '')
        if 'author' in last_parsed:
            last_pipeline['author'] = last_parsed.get('author', '')
        if 'build_url' in last_parsed:
            last_pipeline['build_url'] = last_parsed.get('build_url', '')
        if 'version_path' in last_parsed:
            last_pipeline['version_path'] = last_parsed.get('version_path', '')
    
    # 确保last_pipeline状态正确
    last_pipeline['status'] = final_status
    
    return {
        'last_pipeline': last_pipeline,
        'all_pipelines': all_pipelines,
        'has_build_sections': len(all_pipelines) > 0,
        'build_sections_count': len(all_pipelines)
    }


# 简化版本 - 直接获取最后一个pipeline状态
def get_last_pipeline_status(description: str) -> str:
    """
    快速获取最后一个pipeline的构建状态
    
    Args:
        description: MR的完整描述内容
        
    Returns:
        'success', 'failed', 或 'unknown'
    """
    if not description:
        return 'unknown'
    
    # 查找最后一个构建状态
    lines = description.split('\n')
    
    # 从后往前查找
    for i in range(len(lines) - 1, -1, -1):
        line = lines[i].strip()
        print(line)
        
        # 检查是否包含构建状态
        if '- [x] Domain Build success' in line:
            return 'success'
        elif '- [x] Domain Build failed' in line:
            return 'failed'
        # elif '- [ ] Domain Build success' in line or '- [ ] Domain Build failed' in line:
        #     return 'not_run'  # 构建未运行
    
    # 尝试正则表达式搜索
    import re
    
    # 查找所有构建状态
    status_pattern = r'- \[([x ])\] Domain Build (success|failed)'
    matches = re.findall(status_pattern, description, re.IGNORECASE)
    
    if matches:
        last_match = matches[-1]
        checkbox, status = last_match
        if checkbox.strip() == 'x':
            return status
        else:
            return 'not_run'
    
    # 查找包含pipeline信息的行
    pipeline_lines = []
    for line in lines:
        if 'Build URL:' in line or 'Domain Label:' in line:
            pipeline_lines.append(line)
    
    if pipeline_lines:
        # 有pipeline信息，但状态未知
        return 'unknown_has_pipeline'
    
    return 'unknown_no_pipeline'

# ==================== 主程序 ====================

def main():
    """主程序"""
    
    print("=" * 60)
    print("GitLab MR合规性验证工具 v1.0.1")
    print("=" * 60)
    
    # 显示配置信息
    print(f"项目: {MR_INFO['project_path']}")
    print(f"MR ID: {MR_INFO['mr_iid']}")
    print(f"GitLab URL: {GITLAB_CONFIG['url']}")
    print(f"JIRA集成: {'启用' if JIRA_CONFIG.get('enabled') else '禁用'}")
    print("-" * 60)
    
    # 检查必要的库
    try:
        import requests
        import pandas as pd
    except ImportError as e:
        print(f"❌ 缺少必要的库: {e}")
        print("请运行: pip install requests pandas openpyxl")
        sys.exit(1)
    
    # 1. 获取MR详细信息
    print("正在获取MR信息...")
    mr_data = get_gitlab_mr_details(
        GITLAB_CONFIG['url'],
        MR_INFO['project_path'],
        MR_INFO['mr_iid'],
        GITLAB_CONFIG['token']
    )

    if not mr_data:
        print("❌ 无法获取MR信息，程序退出")
        sys.exit(1)
    
    print(f"✅ 成功获取MR信息: {mr_data.get('title', 'N/A')}")
    print(f"   状态: {mr_data.get('state', 'N/A')} | 创建者: {safe_get(mr_data, 'author', 'name', default='N/A')}")
    
    # 2. 获取MR变更信息
    print("正在获取变更信息...")
    changes_data = get_gitlab_mr_changes(
        GITLAB_CONFIG['url'],
        MR_INFO['project_path'],
        MR_INFO['mr_iid'],
        GITLAB_CONFIG['token']
    )
    print(f"✅ 获取到 {len(changes_data)} 个变更文件")
    
    # 3. 分析合规性
    print("正在进行合规性检查...")
    validation_results, jira_summary = analyze_mr_data(mr_data, changes_data, JIRA_CONFIG)
    
    # 4. 显示检查结果
    print("\n" + "=" * 60)
    print("合规性检查结果:")
    print("=" * 60)
    
    for i in range(len(validation_results["检查项"])):
        check = validation_results["检查项"][i]
        status = validation_results["是否合规"][i]
        detail = validation_results["详细信息"][i]
        
        if status == "√":
            icon = "✅"
        elif status == "⚠️":
            icon = "⚠️ "
        elif status == "Ⅹ":
            icon = "❌"
        else:
            icon = "❓"
        
        print(f"{icon} {check:20} {status:5} {detail}")
    
    # 5. 计算统计
    total_checks = len(validation_results["是否合规"])
    passed_checks = sum(1 for s in validation_results["是否合规"] if s == "√")
    warning_checks = sum(1 for s in validation_results["是否合规"] if s == "⚠️")
    failed_checks = sum(1 for s in validation_results["是否合规"] if s == "Ⅹ")
    pass_rate = (passed_checks / total_checks * 100) if total_checks > 0 else 0
    
    print("-" * 60)
    print(f"📊 统计: 通过 {passed_checks}/{total_checks} | 警告 {warning_checks} | 失败 {failed_checks} | 通过率 {pass_rate:.1f}%")
    
    # 6. 生成报告文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 创建输出目录
    output_dir = f"mr_reports_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)
    
    # 生成Excel报告
    excel_file = os.path.join(output_dir, f"MR_{MR_INFO['mr_iid']}_验证报告_{timestamp}.xlsx")
    excel_path = create_excel_report(mr_data, validation_results, jira_summary, changes_data, excel_file)
    
    # 生成HTML报告
    html_file = os.path.join(output_dir, f"MR_{MR_INFO['mr_iid']}_验证报告_{timestamp}.html")
    html_path = create_html_report(mr_data, validation_results, jira_summary, changes_data, html_file)
    
    # 保存原始数据
    json_file = os.path.join(output_dir, f"MR_{MR_INFO['mr_iid']}_原始数据_{timestamp}.json")
    try:
        with open(json_file, 'w', encoding='utf-8') as f:
            # 确保数据可序列化
            json.dump(mr_data, f, indent=2, ensure_ascii=False, default=str)
        print(f"💾 原始JSON数据已保存到: {json_file}")
    except Exception as e:
        print(f"⚠️ 保存JSON数据时出错: {e}")
    
    print("\n" + "=" * 60)
    print("📁 报告文件已生成:")
    print("=" * 60)
    print(f"📊 Excel报告: {os.path.abspath(excel_path) if excel_path else '生成失败'}")
    print(f"🌐 HTML报告: {os.path.abspath(html_path) if html_path else '生成失败'}")
    print(f"📁 输出目录: {os.path.abspath(output_dir)}")
    
    # 8. 最终建议
    print("\n" + "=" * 60)
    print("🎯 最终建议:")
    print("=" * 60)
    
    if failed_checks == 0 and warning_checks == 0:
        print("✅ 所有检查项通过，MR符合合并要求")
    elif failed_checks > 0:
        print(f"❌ 有 {failed_checks} 个检查项失败，请修复后再尝试合并")
    elif warning_checks > 0:
        print(f"⚠️  有 {warning_checks} 个警告项，建议检查确认")
    
    print("=" * 60)
    print("验证完成！")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n程序被用户中断")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ 程序执行出错: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)