import requests
import json
import sys
import re
from urllib.parse import quote,urljoin
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font,Border, Side
from configparser import ConfigParser


TARGET_BRANCH = "8295_master"  
GROUP_PATH = "RSE"


# === 获取所有以 domain:: 开头的标签 ===
# 得到所有domain label
def get_all_mrs(GITLAB_URL,GROUP_PATH,ACCESS_TOKEN):
    all_mrs_info = {}
    son_mrs = []
    page = 1
    per_page = 100  # 每页最大数量
    
    while True:
        # 构建API请求URL
        url = f"{GITLAB_URL}/api/v4/groups/{GROUP_PATH}/merge_requests"
         # 请求头
        headers = {
            "PRIVATE-TOKEN": ACCESS_TOKEN
        } 
        # 请求参数
        params = {
            "state": "opened",
            "scope": "all",
            "page": page,
            "per_page": per_page,
            "target_branch": TARGET_BRANCH,
            "labels": "ready_to_sys_staging_pending"
        }
        
       
        # 发送GET请求
        response = requests.get(url, params=params, headers=headers)       
        if response.status_code != 200:
            print(f"请求失败，状态码: {response.status_code}")
            print(response.text)
            break
        
        mrs = response.json()
        if not mrs:
            break
            
        # 提取所需信息
        for mr in mrs:
            # 主MR的web_url
            main_url = mr.get("web_url")
            description = mr.get("description", "")
            # 只保留domain::开头的标签
            domain_labels = [label for label in mr.get('labels', []) if label.startswith('domain::')]

            # 提取 Crosslink Merge Requests: 到 Domain: 之间的MR
            match = re.search(r"Crosslink Merge Requests:?\s*(.*?)\s*Domain:", description, re.S)
            if match:
                section = match.group(1)
                # 主MR包含的子MR，去重
                son_urls = list(set(re.findall(r"https://git\.sb\.com[^\s]+/merge_requests/\d+", section)))
            else:
                son_urls = []

            all_mrs_info[main_url] = {
                "son_mrs": son_urls
            }
        
        # 如果没有子MR
        # if len(son_urls) == 0:
        #     print("\033[1;31m  label可能存在问题，请检查！ \033[0m")
        #     print(main_url)
        #     print("\033[1;31m  label可能存在问题，请检查！ \033[0m")
        page += 1
        
        # 检查是否有更多页
        if 'next' not in response.links:
            break
    # 写入符合条件的 MR 到文件
    # with open("filtered_mrs.json", "w", encoding="utf-8") as f:
    #     json.dump(all_mrs_info,f, indent=4, ensure_ascii=False)
    #print(f"总共写入 {len(all_mrs)} 个符合条件的 MR 到 filtered_mrs.json")
    return all_mrs_info


def get_project_id(project_path, ACCESS_TOKEN):
    encoded_path = quote(project_path, safe='')
    url = f"{GITLAB_URL}/api/v4/projects/{encoded_path}"
    
    headers = {"PRIVATE-TOKEN": ACCESS_TOKEN}
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        project_data = response.json()
        return project_data['id']
    else:
        print(f"获取项目ID失败，状态码: {response.status_code}")
        print(f"响应内容: {response.text}")
        return None


def add_label_to_mr(all_mrs_info):
    NEW_LABEL = "ready_to_sys_staging"
    mrs_info = []
   
    for main_url, mr_info in all_mrs_info.items():
        for _, mr_list in mr_info.items():
            for url in mr_list:
                mr_iid = url.split("/")[-1]
                pattern = r'https://git\.sb\.com/(.+?)/-/merge_requests/\d+'
                match = re.match(pattern, url)
                if match:
                    repo_path = match.group(1)
                    # print(f"仓库名称: {repo_path}")
                    project_id = get_project_id(repo_path, ACCESS_TOKEN)
                else:
                    print("未能提取到仓库名称")

                web_url = urljoin(GITLAB_URL, f"api/v4/projects/{project_id}/merge_requests/{mr_iid}")

                headers = {
                    "PRIVATE-TOKEN": ACCESS_TOKEN
                }

                try:
                    response = requests.get(web_url, headers=headers)
                    response.raise_for_status()
                    # 返回mr相关信息
                    mr_data = response.json()
                    source_branch = mr_data.get("source_branch")
                    mr_url = mr_data.get("web_url")
                    current_labels = set(mr_data.get("labels", []))
                    # 对 mr 信息进行备份
                    mr_data = {
                        "web_url" : mr_url,
                        "source_branch": source_branch
                    }
                    
                except requests.exceptions.RequestException as e:
                    print(f"[ERROR] 获取 MR 信息失败 ({url}): {e}")
                    continue

                if NEW_LABEL in current_labels:
                    print(f"[SKIP] MR 已包含标签 '{NEW_LABEL}'，跳过: {url}")
                    continue

                updated_labels = current_labels.union({NEW_LABEL})
                label_str = ",".join(updated_labels)
                
                try:
                    update_data = {"labels": label_str}
                    response = requests.put(web_url, headers=headers, data=update_data)
                    response.raise_for_status()
                    print(f"[SUCCESS] 成功为 {url} 添加标签: {NEW_LABEL}")
                except requests.exceptions.RequestException as e:
                    print(f"[ERROR] 更新 MR 标签失败 ({url}): {e}")
                
                mrs_info.append(mr_data)

    # 写入符合条件的 MR 到文件
    with open("mr_info.json", "w", encoding="utf-8") as f:
        json.dump(mrs_info,f, indent=4, ensure_ascii=False)
    

def set_center_cell(ws, row, col, value,border):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
    return cell

def write_excel(all_mrs_info):
    wb = Workbook()
    ws = wb.active
    ws.title = "MR Info"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )


    # 表头设置
    headers = ["MR COUNT", "MAIN URL", "SON URL", "PROJECT ID", "MR IID"]
    header_font = Font(bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header).border = thin_border
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 2  # 从第二行开始写数据

    for main_url, info in all_mrs_info.items():
        mr_count = len(info["son_mrs"])
        start_row = current_row

        for son in info["son_mrs"]:
            set_center_cell(ws, current_row, 1, mr_count,thin_border)
            ws.cell(row=current_row, column=2, value=main_url).border = thin_border
            ws.cell(row=current_row, column=3, value=son).border = thin_border
            # set_center_cell(ws, current_row, 4, son["project_id"],thin_border)
            # set_center_cell(ws, current_row, 5, son["mr_iid"],thin_border)
            set_center_cell(ws, current_row, 4, "project_id",thin_border)
            set_center_cell(ws, current_row, 5, "mr_iid",thin_border)
            current_row += 1

        end_row = current_row - 1

        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)

    # 冻结表头（可选）
    ws.freeze_panes = "A2"
    # 设置列宽（可选）
    col_widths = [10, 40, 40, 15, 15]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    wb.save("mrs.xlsx")
    print("Excel 文件已生成：mrs.xlsx")


if __name__ == "__main__":
    config = ConfigParser()
    config.read("/home/gayuxia/config.ini")
    ACCESS_TOKEN = config.get("gitlab","token")
    GITLAB_URL = config.get("gitlab","url")
    # === 配置信息 ===
    # 这里会得到去掉domain::后的label
    all_mrs_info = get_all_mrs(GITLAB_URL, GROUP_PATH, ACCESS_TOKEN)    
    add_label_to_mr(all_mrs_info)
 
    write_excel(all_mrs_info)
    print("\033[1;31m 请手动检查带有 asset_translation_mr or assets_translation_mr & ready_to_sys_staging_pending 标签的MR \033[0m") 
    print("\033[1;31m 请手动检查带有 asset_translation_mr or assets_translation_mr & ready_to_sys_staging_pending 标签的MR \033[0m") 
    print("\033[1;32m 如果遇到删除的MR，需要检查它在后面的 review 过程中有没有加回来 \033[0m") 
    print("\033[1;32m 如果遇到删除的MR，需要检查它在后面的 review 过程中有没有加回来 \033[0m") 

