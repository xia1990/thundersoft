import requests
import json
import sys
import re
from urllib.parse import quote
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from urllib.parse import quote
from common.Query import Query
from configparser import ConfigParser


def get_mr_info(GITLAB_URL, project_id, mr_iid, ACCESS_TOKEN):
    """获取指定 MR 的 commitID 和 jira_id"""
    url = f"{GITLAB_URL}/api/v4/projects/{project_id}/merge_requests/{mr_iid}"
    headers = {"PRIVATE-TOKEN": ACCESS_TOKEN}
    response = requests.get(url, headers=headers)

    if response.status_code != 200:
        print(f"获取 MR 失败，状态码: {response.status_code}, iid={mr_iid}")
        return {}

    mr = response.json()

    # commitID: 获取 MR 最新 commit
    commits_url = f"{url}/commits"
    commits_resp = requests.get(commits_url, headers=headers)
    commit_id = None
    if commits_resp.status_code == 200 and commits_resp.json():
        commit_id = commits_resp.json()[0].get("id")

    # jira_id: 从 title 中解析
    title = mr.get("title", "")
    #APRICOTBSC
    # match = re.match(r"^(APRICOT-\d+)", title)
    match = re.match(r"^(APRICOT-\d+|APRICOTBSC-\d+)", title)
    jira_id = match.group(1) if match else ""
    label = mr.get("labels")

    return {
        "commitID": commit_id,
        "jira_id": jira_id,
        "web_url": mr.get("web_url"),
        "label": label
    }


def get_domain_labels(GITLAB_URL, GROUP_PATH, ACCESS_TOKEN):
    all_mrs_info = {}
    main_info = {}
    page = 1
    per_page = 100

    while True:
        url = f"{GITLAB_URL}/api/v4/groups/{GROUP_PATH}/merge_requests"
        headers = {"PRIVATE-TOKEN": ACCESS_TOKEN}
        params = {
            "state": "opened",
            "scope": "all",
            "page": page,
            "per_page": per_page,
            "target_branch": TARGET_BRANCH,
            "labels": "ready_to_review,domain::*"
            # "labels": "ready_to_sys_staging_pending"
        }

        response = requests.get(url, params=params, headers=headers)
        if response.status_code != 200:
            print(f"请求失败，状态码: {response.status_code}")
            print(response.text)
            break

        mrs = response.json()
        if not mrs:
            break

        for mr in mrs:
            main_url = mr.get("web_url")
            main_label = mr.get("labels")
            project_id = mr.get("project_id")
            mr_iid = mr.get("iid")
            description = mr.get("description") or ""
            match = re.search(r'Domain:\s*([^\n,]+)', description)
            domain = match.group(1).strip() if match else None
            #
            main_info[mr_iid] = {
                "mr_iid": mr_iid,
                "project_id": project_id,
                "web_url": main_url
            }
            # 获取主 MR 的信息
            main_mr_info = get_mr_info(GITLAB_URL, project_id, mr_iid, ACCESS_TOKEN)

            # 提取 Crosslink Merge Requests 部分的子 MR
            match = re.search(r"Crosslink Merge Requests:?\s*(.*?)\s*Domain:", description, re.S)
            if match:
                section = match.group(1)
                son_urls_source = list(set(re.findall(r"https://git\.sb\.com[^\s]+/merge_requests/\d+", section)))
            else:
                son_urls_source = []

            # 子 MR 信息
            son_mrs_info = []
            for son_url in son_urls_source:
                # 提取 project_path 和 iid
                son_match = re.match(r"https://git\.sb\.com/(.+?)/-/merge_requests/(\d+)", son_url)
                if not son_match:
                    continue
                son_project_path, son_iid = son_match.groups()

                # 获取 project_id
                proj_resp = requests.get(
                    f"{GITLAB_URL}/api/v4/projects/{requests.utils.quote(son_project_path, safe='')}",
                    headers=headers
                )
                if proj_resp.status_code != 200:
                    print(f"获取子项目失败: {son_project_path}")
                    continue
                son_project_id = proj_resp.json().get("id")

                # 获取子 MR 详细信息
                son_info = get_mr_info(GITLAB_URL, son_project_id, son_iid, ACCESS_TOKEN)
                if son_info:
                    son_mrs_info.append(son_info)

            all_mrs_info[main_url] = {
                "main_label": main_label,
                "main_url": domain,
                "son_mrs": son_mrs_info
            }

        page += 1
        if 'next' not in response.links:
            break
    # 将得到的信息写入 json 文件中
    with open("need_to_review_mrs_info.json", "w", encoding="utf-8") as f:
        json.dump(all_mrs_info, f, indent=4, ensure_ascii=False)

    return all_mrs_info,main_info


def write_mrs_to_excel(all_mrs_info, output_file="mr_review_report.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MR Review Report"

    # 调整headers顺序，将Status放到第3列
    headers = [
        "Domain", "No.", "Staging Review", "Primary MR", "MRs",
        "Ticket", "Revision", "Review Result", "Comment"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    MR_TOTAL = 0
    row_num = 2
    for main_url, mr_info in all_mrs_info.items():
        domain = mr_info.get("main_url", "")
        son_urls = mr_info.get("son_mrs", [])
        labels = mr_info.get("main_label")
        count = len(son_urls)
        mr_type = "Multi MR" if count > 1 else "Manifest MR"
        ready_to_review = "Yes" if "ready_to_review" in labels else "No"

        merge_rows = max(1, count)
        # 合并前 4 列
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num+merge_rows-1, end_column=1)  # Domain
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num+merge_rows-1, end_column=2)  # Number
        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num+merge_rows-1, end_column=3)  # Status
        ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num+merge_rows-1, end_column=4)  # Multi MR or Manifest MR

        # 填充合并单元格的数据
        ws.cell(row=row_num, column=1, value=domain)  # Domain
        ws.cell(row=row_num, column=2, value=count)   # Number
        ws.cell(row=row_num, column=3, value="✅" if "ready_ng" not in labels else "❌")  # Status
        ws.cell(row=row_num, column=4, value=main_url)  # Multi MR or Manifest MR

        # 得到所有MR的数量
        MR_TOTAL += count
        # 填充子MR的详细信息
        for i, son in enumerate(son_urls):
            current_row = row_num + i
            ws.cell(row=current_row, column=5, value=son.get("web_url", "")).alignment = Alignment(wrap_text=True)  # MRs
            jira_id = son.get("jira_id", "")
            
            # 初始化第15列的值
            review_result = ""  # 默认空值
            
            # 判断 jira Ticket 中是否包含 E241.0_MUSTFIX
            if jira_id:
                jira_url = f"https://issue.sb.com/rest/api/2/issue/{jira_id}"  # 注意：这里应该是API URL，不是browse URL
                headers = {
                    "Authorization": f"Bearer {API_TOKEN}",
                    "Content-Type": "application/json"
                }
                try:
                    response = requests.get(jira_url, headers=headers, timeout=10)
                    # 解析响应
                    if response.status_code == 200:
                        issue_data = response.json()
                        jira_labels = issue_data.get("fields", {}).get("labels", [])
                        
                        # 检查是否包含 E241.0_MUSTFIX
                        if LABEL in jira_labels:
                            print(f"{son.get('web_url', '')} ✅ Label {LABEL} found!")
                            review_result = f"{LABEL}"  # 在第15列写上这个label
                        else:
                            print(f"{jira_id} ❌ Label {LABEL} not found.")
                    else:
                        print(f"❌ Failed to fetch issue {jira_id}: {response.status_code} - {response.text}")
                except Exception as e:
                    print(f"❌ Error fetching JIRA data for {jira_id}: {e}")
            
            # 设置JIRA Ticket超链接
            if jira_id:
                jira_link = f'=HYPERLINK("https://issue.sb.com/browse/{jira_id}", "{jira_id}")'
                ws.cell(row=current_row, column=6, value=jira_link).alignment = Alignment(wrap_text=True)  # Ticket
            else:
                ws.cell(row=current_row, column=6, value="").alignment = Alignment(wrap_text=True)  # Ticket为空
            
            ws.cell(row=current_row, column=7, value=son.get("commitID", "")).alignment = Alignment(wrap_text=True)  # Revision
            ws.cell(row=current_row, column=8, value="")  # Review Result 
            # 第12-14列暂时为空
            ws.cell(row=current_row, column=9, value="")  # Comment

        row_num += merge_rows

    # 调整列宽（对应新的列顺序）
    widths = [10, 5, 5, 80, 80, 20, 15, 15, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(output_file)
    print(f"报告已生成，保存为: {output_file}")


def remove_label_from_gitlab_mr(GITLAB_URL, GROUP_PATH, ACCESS_TOKEN, main_info):
    remove_label = "ready_to_review"
    for mr_data in main_info.values():
        project_id = mr_data.get("project_id")
        mr_iid = mr_data.get("mr_iid")
        url = f"{GITLAB_URL}/api/v4/projects/{project_id}/merge_requests/{mr_iid}"
    
        headers = {
            "PRIVATE-TOKEN": ACCESS_TOKEN
        }
    
        # 首先获取当前标签
        response = requests.get(url, headers=headers)
        current_labels = response.json().get("labels", [])
    
        if remove_label in current_labels:
            updated_labels = [label for label in current_labels if label != remove_label]
        else:
            updated_labels = current_labels

        # 发送更新请求
        update_resp = requests.put(url, headers=headers, json={"labels": updated_labels})
        if update_resp.status_code == 200:
            print(f"已从 MR {mr_iid} 中删除标签: {remove_label}")
        else:
            print(f"删除失败: {update_resp.status_code}, {update_resp.text}")


# === 主函数 ===
def main():
    try:
        # 这里会得到去掉domain::后的label
        all_mrs_info,main_info = get_domain_labels(GITLAB_URL, GROUP_PATH, ACCESS_TOKEN)
        write_mrs_to_excel(all_mrs_info, output_file="mr_review_report.xlsx")
        # remove_label_from_gitlab_mr(GITLAB_URL, GROUP_PATH, ACCESS_TOKEN, main_info)
    except Exception as e:
        print(f"发生错误: {e}")


if __name__ == "__main__":
    config = ConfigParser()
    config.read("/home/gayuxia/config.ini")
    ACCESS_TOKEN = config.get("gitlab","token")
    GITLAB_URL = config.get("gitlab","url")

    GROUP_PATH = "RSE"
    TARGET_BRANCH = "8295_master" 
    LABEL = "E247.0_MUSTFIX"
    API_TOKEN = "token"
    # === 主函数 ======
    main()
