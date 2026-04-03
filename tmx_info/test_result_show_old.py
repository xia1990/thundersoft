#!/usr/bin/python3

from openpyxl import Workbook
from openpyxl.styles import Alignment
import requests
import json
from configparser import ConfigParser



wb = Workbook()
ws1 = wb.active
ws1.title = "Overall(STAR3.0)"
ws2 = wb.create_sheet(title="Overall(STAR3.5)", index=1) 
ws3 = wb.create_sheet(title="Overall(RSU VIP)", index=1) 

# JIRA info
JIRA_URL = "https://issue.sb.com"
API_TOKEN = "token"  


jql_query = """
project = TMX AND issuetype = "Test Execution" AND assignee in (shaoli) ORDER BY summary ASC, priority DESC, updated DESC
"""

def get_jira_issues(jql):
    all_issues = []
    start_at = 0
    max_results = 100  # 每页最多 100 条

    while True:
        url = f"{JIRA_URL}/rest/api/2/search"

        headers = {
            "Authorization": f"Bearer {API_TOKEN}",
            "Content-Type": "application/json"
        }

        params = {
            "jql": jql,
            "startAt": start_at,
            "maxResults": max_results,
            "fields": "key,summary,status,assignee,priority,updated" 
        }
        
        response = requests.get(
            url,
            headers=headers,
            params=params
        )
        
        if response.status_code != 200:
            print(f"请求失败！状态码: {response.status_code}")
            # print(response.text)
            break

        data = response.json()
        issues = data.get("issues", [])
        all_issues.extend(issues)
        
        # 如果已获取全部结果，退出循环
        if len(issues) < max_results:
            break
            
        start_at += max_results
    
    return all_issues


def get_tmx_info(issues):
    all_tmx_info = {}
    headers = {
        "Authorization": f"Bearer {API_TOKEN}",
        "Content-Type": "application/json"
    }

    for issue in issues:
        summary = issue["fields"]["summary"]
        version_type = summary.split("_")[0].strip()
        
        # 所有版本类型的处理逻辑相同，无需重复代码
        jira_key = issue["key"]
        
        try:
            response = requests.get(
                f"{JIRA_URL}/rest/api/2/issue/{jira_key}",
                headers=headers
            )
            response.raise_for_status()  # 检查HTTP错误
            data = response.json()
            
            # 提取测试状态信息
            test_info = extract_test_status_info(data)
            
            if test_info:
                all_tmx_info[jira_key] = test_info
                
        except requests.exceptions.RequestException as e:
            print(f"Error fetching data for {jira_key}: {e}")
        except (KeyError, TypeError) as e:
            print(f"Error parsing data for {jira_key}: {e}")
    
    return all_tmx_info


def extract_test_status_info(data):
    """从JIRA响应数据中提取测试状态信息"""
    passed = failed = nt = total = 0
    blocked_found = False
    
    # 递归搜索包含customfield_11120的字典
    def find_customfield(obj):
        if isinstance(obj, dict):
            if "customfield_11120" in obj:
                return obj["customfield_11120"]
            for value in obj.values():
                result = find_customfield(value)
                if result:
                    return result
        elif isinstance(obj, list):
            for item in obj:
                result = find_customfield(item)
                if result:
                    return result
        return None
    
    custome = find_customfield(data)
    
    if not custome:
        return None
    
    # 提取测试统计信息
    total = custome.get("count", 0)
    statuses = custome.get("statuses", [])
    
    for status in statuses:
        name = status.get("name")
        count = status.get("statusCount", 0)
        
        if name == "PASS":
            passed = count
        elif name == "FAIL":
            failed = count
        elif name == "BLOCKED":
            nt = count
            if count != 0:
                blocked_found = True
        elif name == "TODO" and not blocked_found:
            nt = count
    
    return {
        "total": total,
        "passed": passed,
        "failed": failed,
        "nt": nt
    }
        

def query_mr(mr_label,GITLAB_URL,GROUP_PATH,ACCESS_TOKEN):
    all_mrs=[]
    new_label = f"integrated_{mr_label}"
    page = 1
    per_page = 100  # 每页最大数量
    
    # 构建API请求URL
    url = f"{GITLAB_URL}/api/v4/groups/{GROUP_PATH}/merge_requests"
    # 请求头
    headers = {
        "PRIVATE-TOKEN": ACCESS_TOKEN
    } 
    # 请求参数
    # 请求参数
    params = {
        "scope": "all",
        "page": page,
        "per_page": per_page,
        "labels": new_label
    }
        
    # 发送GET请求
    response = requests.get(url, params=params, headers=headers)       
    if response.status_code != 200:
        print(f"请求失败，状态码: {response.status_code}")
        
    mrs = response.json()
    
    for mr in mrs:
        mr_iid = mr.get("iid")
        all_mrs.append(mr_iid)
    return all_mrs


def write_excel(issues, tmx_info):
    # 定义工作表配置
    sheet_configs = {
        "STAR3.0": {
            "ws": ws1,
            "headers": ["i3 RSU Star 3.0", "Passed", "Blocked", "Failed", "NT", "Total", "MR Number"],
            "col_counter": 1
        },
        "STAR3.5": {
            "ws": ws2,
            "headers": ["i3 RSU Star 3.5", "Passed", "Blocked", "Failed", "NT", "Total", "MR Number"],
            "col_counter": 1
        },
        "default": {
            "ws": ws3,
            "headers": ["RSU VIP", "Passed", "Blocked", "Failed", "NT", "Total", "MR Number"],
            "col_counter": 1
        }
    }

    # 先写入所有表头
    for config in sheet_configs.values():
        for header in config["headers"]:
            config["ws"].append([header])

    # 预处理MR查询结果，避免重复查询
    mr_cache = {}
    for issue in issues:
        summary = issue["fields"]["summary"]
        parts = summary.split("_")
        if len(parts) >= 2:
            mr_label = parts[1]
            if mr_label not in mr_cache:
                mr_cache[mr_label] = len(query_mr(mr_label, GITLAB_URL, GROUP_PATH, ACCESS_TOKEN))

    # 按issue处理数据，避免嵌套循环
    for issue in issues:
        key = issue["key"]
        if key not in tmx_info:
            continue
            
        summary = issue["fields"]["summary"]
        parts = summary.split("_")
        
        if len(parts) < 3:
            continue
            
        version_type = parts[0].strip()
        build_date = parts[-1]
        build_type = "_".join(parts[1:3])
        mr_label = parts[1]
        
        data = tmx_info[key]
        mr_number = mr_cache.get(mr_label, 0)
        
        # 确定使用哪个工作表配置
        config_key = version_type if version_type in sheet_configs else "default"
        config = sheet_configs[config_key]
        
        # 更新列计数器
        config["col_counter"] += 1
        col_index = config["col_counter"]
        
        # 写入数据
        ws = config["ws"]
        cell_value = f"{build_date}\n{build_type}"
        
        # 写入单元格数据
        ws.cell(row=1, column=col_index, value=cell_value)
        ws.cell(row=1, column=col_index).alignment = Alignment(wrap_text=True)
        ws.cell(row=2, column=col_index, value=data.get("passed", 0))
        ws.cell(row=3, column=col_index, value=0)  # Blocked固定为0
        ws.cell(row=4, column=col_index, value=data.get("failed", 0))
        ws.cell(row=5, column=col_index, value=data.get("nt", 0))
        ws.cell(row=6, column=col_index, value=data.get("total", 0))
        ws.cell(row=7, column=col_index, value=mr_number)

    # 保存 Excel 文件
    wb.save("rsu_staging_trend.xlsx")
    print("Excel 文件已生成：rsu_staging_trend.xlsx")


if __name__ == '__main__':
    config = ConfigParser()
    config.read("/home/gaoyuxia/config.ini")
    ACCESS_TOKEN = config.get("gitlab","token")
    GITLAB_URL = config.get("gitlab","url")
    GROUP_PATH = "RSE"

    issues = get_jira_issues(jql_query)
    # print(f"共找到 {len(issues)} 个问题!")
    tmx_info = get_tmx_info(issues)
    write_excel(issues,tmx_info) 
    