from configparser import ConfigParser
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def query_mr(mr_label,GITLAB_URL,GROUP_PATH,ACCESS_TOKEN):
    all_mrs=[]
    new_label = f"integrated_{mr_label}"
    page = 1
    per_page = 100  # 每页最大数量
    
    # 构建API请求URL
    url = f"{GITLAB_URL}/api/v4/groups/{GROUP_PATH}/merge_requests"
    # 请求头
    headers = {
        "PRIVATE-TOKEN": ACCESS_TOKEN
    } 
    # 请求参数
    params = {
        "state": "merged",
        "scope": "all",
        "page": page,
        "per_page": per_page,
        "labels": mr_label
    }
        
    # 发送GET请求
    response = requests.get(url, params=params, headers=headers)      
    if response.status_code != 200:
        print(f"请求失败，状态码: {response.status_code}")
        
    mrs = response.json()
    for mr in mrs:
        mr_iid = mr.get("iid")
        all_mrs.append(mr_iid)
    return all_mrs


def write_to_excel_pandas(version_total, mr_labels, GITLAB_URL, GROUP_PATH, ACCESS_TOKEN):
    # 准备数据
    data = []
    for label in mr_labels:
        all_mrs = query_mr(label, GITLAB_URL, GROUP_PATH, ACCESS_TOKEN)
        mr_totals = len(all_mrs)
        version = label.replace("integrated_", "")
        
        data.append({
            'Version Total': version_total,
            'Version': version,
            'MR Totals': mr_totals
        })
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 先写入Excel文件（不合并单元格）
    filename = "mr_statistics.xlsx"
    df.to_excel(filename, index=False)
    
    # 使用openpyxl加载文件并合并单元格
    wb = load_workbook(filename)
    ws = wb.active
    
    # 合并Version Total列的单元格（从第2行到第2+len(mr_labels)-1行）
    start_row = 2
    end_row = start_row + len(mr_labels) - 1
    
    if end_row > start_row:  # 只有多行数据时才需要合并
        merge_range = f'A{start_row}:A{end_row}'
        ws.merge_cells(merge_range)
        
        # 设置合并后单元格的垂直居中
        merged_cell = ws[f'A{start_row}']
        merged_cell.alignment = Alignment(vertical='center', horizontal='center')
    
    # 保存修改后的文件
    wb.save(filename)
    print(f"数据已成功写入 {filename}，Version Total 列已合并")


if __name__ == '__main__':
    config = ConfigParser()
    config.read("/home/GAYUXIA/config.ini")
    ACCESS_TOKEN = config.get("gitlab","token")
    GITLAB_URL = config.get("gitlab","url")
    GROUP_PATH = "RSE"

    mr_labels = ["integrated_E236.0-35.6","integrated_E236.0-36.2","integrated_E236.0-36.3","integrated_E236.0-36.6",
    "integrated_E237.0-36.6","integrated_E237.0-37.2","integrated_E237.0-37.3","integrated_E237.0-37.4","integrated_E237.0-38.2","integrated_E237.0-38.3",
    "integrated_E237.0-38.6","integrated_E238.0-39.3","integrated_E238.0-39.5"]
    version_total = len(mr_labels)
    for label in mr_labels:
        all_mrs = query_mr(label,GITLAB_URL,GROUP_PATH,ACCESS_TOKEN)
        mr_totals = len(all_mrs)
        version = label.replace("integrated_","")
    write_to_excel_pandas(version_total, mr_labels, GITLAB_URL, GROUP_PATH, ACCESS_TOKEN)   
        