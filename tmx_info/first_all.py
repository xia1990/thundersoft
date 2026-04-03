#!/usr/bin/python3
# -*- coding: utf-8 -*-

import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from configparser import ConfigParser
import concurrent.futures
import sys
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Side
import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

# ========== 初始化 Excel ==========
wb = Workbook()
ws1 = wb.active
ws1.title = "Star3.0"
ws2 = wb.create_sheet(title="Star3.5", index=1)

# ========== JIRA 查询 ==========
sql_query = """
project = TMX AND issuetype = "Test Execution" AND reporter in (shaoli) ORDER BY priority DESC, updated DESC
"""

def get_jira_issues(session, jql):
    """分页获取所有 JIRA issues"""
    all_issues = []
    start_at = 0
    max_results = 100

    while True:
        url = f"{JIRA_URL}/rest/api/2/search"
        params = {
            "jql": jql,
            "startAt": start_at,
            "maxResults": max_results,
            "fields": "key,summary,status,assignee,priority,updated"
        }

        response = session.get(url, params=params)
        if response.status_code != 200:
            print(f"❌ 获取 issue 列表失败：{response.status_code}")
            break

        data = response.json()
        issues = data.get("issues", [])
        all_issues.extend(issues)
        if len(issues) < max_results:
            break
        start_at += max_results

    return all_issues


# ========== 并发获取 JIRA 详情 ==========
def extract_test_status_info(data):
    """提取测试统计信息"""
    def find_customfield(obj):
        if isinstance(obj, dict):
            if "customfield_11120" in obj:
                return obj["customfield_11120"]
            for v in obj.values():
                r = find_customfield(v)
                if r:
                    return r
        elif isinstance(obj, list):
            for item in obj:
                r = find_customfield(item)
                if r:
                    return r
        return None

    custome = find_customfield(data)
    if not custome:
        return None

    passed = 0
    total = custome.get("count", 0)
    statuses = custome.get("statuses", [])
    for s in statuses:
        if s.get("name") == "PASS":
            passed = s.get("statusCount", 0)

    return {"total": total, "passed": passed}


def fetch_issue_detail(issue, session):
    """单个 issue 请求详情"""
    key = issue["key"]
    try:
        resp = session.get(f"{JIRA_URL}/rest/api/2/issue/{key}")
        resp.raise_for_status()
        data = resp.json()
        test_info = extract_test_status_info(data)
        return key, test_info
    except Exception as e:
        print(f"⚠️ 获取 {key} 失败: {e}")
        return key, None


def get_tmx_info(issues, session):
    """并发获取所有 issue 的测试状态"""
    tmx_info = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_issue_detail, issue, session): issue for issue in issues}
        for f in concurrent.futures.as_completed(futures):
            key, info = f.result()
            if info:
                tmx_info[key] = info
    return tmx_info


# ========== 并发查询 GitLab MR ==========
def query_mr(session, mr_label):
    new_label = f"integrated_{mr_label}"
    url = f"{GITLAB_URL}/api/v4/groups/{GROUP_PATH}/merge_requests"
    params = {"scope": "all", "labels": new_label, "per_page": 100}
    resp = session.get(url, params=params)
    if resp.status_code != 200:
        print(f"⚠️ MR 查询失败({mr_label})：{resp.status_code}")
        return mr_label, 0
    mrs = resp.json()
    return mr_label, len(mrs)


def get_all_mr_counts(mr_labels):
    mr_cache = {}
    with requests.Session() as session:
        session.headers.update({"PRIVATE-TOKEN": ACCESS_TOKEN})
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            results = list(executor.map(lambda x: query_mr(session, x), mr_labels))
        mr_cache = dict(results)
    return mr_cache


# ========== 写入 Excel ==========
def write_excel(issues, tmx_info):
    sheet_configs = {
        "STAR3.0": {"ws": ws1, "headers": ["i3 RSU Star 3.0 Dev", "Test cases passed", "Test cases total", "MR submitted", "MR reviewed", "MR merged", "passed rate"], "col_counter": 1},
        "STAR3.5": {"ws": ws2, "headers": ["i3 RSU Star 3.5 Dev", "Test cases passed", "Test cases total", "MR submitted", "MR reviewed", "MR merged", "passed rate"], "col_counter": 1},
    }

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # 表头
    for config in sheet_configs.values():
        ws = config["ws"]
        for i, header in enumerate(config["headers"], start=1):
            cell = ws.cell(row=i, column=1, value=header)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 收集 MR label
    mr_labels = list({issue["fields"]["summary"].split("_")[1] for issue in issues if "_" in issue["fields"]["summary"]})
    mr_cache = get_all_mr_counts(mr_labels)

    # 汇总数据
    summary_data = {}
    for issue in issues:
        key = issue["key"]
        if key not in tmx_info:
            continue
        summary = issue["fields"]["summary"]
        parts = summary.split("_")
        if len(parts) < 3:
            continue
        version_type = parts[0].strip()
        build_type = "_".join(parts[1:3])
        mr_label = parts[1]
        mr_number = mr_cache.get(mr_label, 0)
        if not build_type.endswith("_DEV") and not build_type.endswith("_dev") and not build_type.endswith("_HQ"):
            continue

        base_version = build_type.rsplit("_", 1)[0]
        data = tmx_info[key]
        summary_data.setdefault(version_type, {}).setdefault(base_version, {"passed": 0, "total": 0, "mr_number": mr_number, "dev_build_type": f"{base_version}_DEV"})
        summary_data[version_type][base_version]["passed"] += data.get("passed", 0)
        summary_data[version_type][base_version]["total"] += data.get("total", 0)

    # 排序 & 写入
    for version_type, versions in summary_data.items():
        if version_type not in sheet_configs:
            continue
        config = sheet_configs[version_type]
        ws = config["ws"]
        for base_version, data in sorted(versions.items(), reverse=True):
            config["col_counter"] += 1
            col_index = config["col_counter"]
            ws.cell(row=1, column=col_index, value=data["dev_build_type"])
            ws.cell(row=2, column=col_index, value=data["passed"])
            ws.cell(row=3, column=col_index, value=data["total"])
            ws.cell(row=6, column=col_index, value=data["mr_number"])
            rate = f"{(data['passed'] / data['total'] * 100):.2f}%" if data["total"] else "0%"
            ws.cell(row=7, column=col_index, value=rate)
            for r in range(1, 8):
                c = ws.cell(row=r, column=col_index)
                c.border = thin_border
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    file_path = "rsu_staging_trend.xlsx"
    wb.save(file_path)
    print(f"✅ Excel 文件已生成并优化：{file_path}")
    # 绘制并插入堆积图
    draw_and_insert_stacked_chart(file_path)


# 绘制堆积柱状图
def draw_and_insert_stacked_chart(excel_path):
    """
    绘制堆积柱状图 + 3条MR折线，并插入Excel的A10
    智能自适应图表宽高、柱宽、字体大小，保证在Excel中清晰可读
    """
    wb = openpyxl.load_workbook(excel_path)
    plt.rcParams['axes.unicode_minus'] = False  # 负号显示为普通减号

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Star"):
            continue
        ws = wb[sheet_name]

        # ===== 读取数据 =====
        build_types, passed_cases, total_cases = [], [], []
        mr_submitted, mr_reviewed, mr_merged = [], [], []

        for col in range(2, ws.max_column + 1):
            # build_type = ws.cell(row=1, column=col).value
            build_type = ws.cell(row=1, column=col).value.replace("_DEV","")
            passed = ws.cell(row=2, column=col).value or 0
            total = ws.cell(row=3, column=col).value or 0
            submitted = ws.cell(row=4, column=col).value or 0
            reviewed = ws.cell(row=5, column=col).value or 0
            merged = ws.cell(row=6, column=col).value or 0

            if build_type and total:
                build_types.append(str(build_type))
                passed_cases.append(passed)
                total_cases.append(total)
                mr_submitted.append(submitted)
                mr_reviewed.append(reviewed)
                mr_merged.append(merged)

        if not build_types:
            continue

        failed_cases = [t - p for t, p in zip(total_cases, passed_cases)]
        passed_rates = [(p / t * 100 if t else 0) for p, t in zip(passed_cases, total_cases)]

        # ===== 智能自适应参数计算 =====
        n = len(build_types)
        max_label_len = max(len(bt) for bt in build_types)

        # 图宽 = 每个柱子基础宽度 + 间距，保证文字不重叠
        base_bar_width = 0.6
        spacing = 0.2  # 柱间基础间距
        bar_width = base_bar_width
        fig_width = max(12, n * (bar_width + spacing))

        # 图高根据最大文字长度适当增加
        fig_height = max(6, max_label_len * 0.15 + 4)

        # 横坐标字体大小和通过率字体大小
        xtick_fontsize = 8 if n <= 20 else max(6, int(160 / n))  # 列数多字体小
        label_fontsize = xtick_fontsize

        # x坐标位置
        x = np.arange(n)

        # ===== 绘图 =====
        fig, ax1 = plt.subplots(figsize=(fig_width, fig_height), dpi=200)
        fig.patch.set_facecolor("#2b2b2b")
        ax1.set_facecolor("#2b2b2b")

        # 堆积柱状图
        ax1.bar(x, passed_cases, width=bar_width, label='Passed', color='#8BC34A', edgecolor='white', linewidth=0.8)
        ax1.bar(x, failed_cases, width=bar_width, bottom=passed_cases, label='Total', color='#00B0F0', edgecolor='white', linewidth=0.8)

        # 三条 MR 折线
        ax1.plot(x, mr_submitted, color='#FFFF00', linewidth=2, marker='o', label='MR Submitted')
        ax1.plot(x, mr_reviewed, color='#7B68EE', linewidth=2, marker='o', label='MR Reviewed')
        ax1.plot(x, mr_merged, color='#FF0000', linewidth=2, marker='o', label='MR Merged')

        # 文本标注（通过率）
        for i, rate in enumerate(passed_rates):
            ax1.text(x[i], total_cases[i] + max(total_cases)*0.02,
                     f"{rate:.2f}%", ha='center', va='bottom',
                     fontsize=label_fontsize, color='white', fontweight='bold')

        # 坐标样式
        ax1.set_xticks(x)
        # 根据最长文字长度和列数，决定是否旋转文字
        # rotation_angle = 0 if max_label_len <= 8 and n <= 20 else 25
        rotation_angle = 0 
        ax1.set_xticklabels(build_types, rotation=rotation_angle, ha='center', fontsize=xtick_fontsize, color='white')
        ax1.tick_params(axis='y', colors='white')
        ax1.grid(axis='y', linestyle='--', alpha=0.3, color='white')

        # 标题与图例
        ax1.set_title(f"RSU {sheet_name} DEV MRQG Test", fontsize=16, fontweight='bold', color='white', pad=15)
        handles, labels = ax1.get_legend_handles_labels()
        ax1.legend(
            handles, labels,
            loc='upper center',
            bbox_to_anchor=(0.5, -0.08),
            ncol=5,
            frameon=False,
            fontsize=label_fontsize,
            labelcolor='white'
        )
        plt.tight_layout(pad=2.0)

        # ===== 保存并插入Excel =====
        chart_file = f"{sheet_name}_stacked_mr_chart.png"
        plt.savefig(chart_file, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close()

        # 删除旧图避免叠加
        ws._images = []
        img = Image(chart_file)
        # 插入 Excel 时保持清晰，宽高按像素设置
        img.width, img.height = int(fig_width * 120), int(fig_height * 120)
        ws.add_image(img, "A10")

    wb.save(excel_path)
    print(f"✅ 堆积柱 + 3条 MR 折线图 已插入到 {excel_path} 的各 sheet A10 位置")


# ========== 主程序 ==========
if __name__ == "__main__":
    config = ConfigParser()
    config.read("/home/gaoyuxia/config.ini")
    ACCESS_TOKEN = config.get("gitlab", "token")
    GITLAB_URL = config.get("gitlab", "url")
    GROUP_PATH = "RSE"
    JIRA_URL = config.get("jira", "JIRA_URL")
    API_TOKEN = config.get("jira", "API_TOKEN")

    with requests.Session() as session:
        session.headers.update({"Authorization": f"Bearer {API_TOKEN}", "Content-Type": "application/json"})
        issues = get_jira_issues(session, sql_query)
        print(f"共找到 {len(issues)} 个 issues")
        tmx_info = get_tmx_info(issues, session)

    write_excel(issues, tmx_info) 
