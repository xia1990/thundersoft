#!/usr/bin/python3

import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment
import requests
import json
import os
from configparser import ConfigParser



jql_query = """
project = TMX AND issuetype = "Test Execution" AND assignee in (shaoli) ORDER BY summary ASC, priority DESC, updated DESC
"""

def get_jira_issues(jql):
    all_issues = []
    start_at = 0
    max_results = 100  # 每页最多 100 条

    while True:
        url = f"{JIRA_URL}/rest/api/2/search"

        headers = {
            "Authorization": f"Bearer {API_TOKEN}",
            "Content-Type": "application/json"
        }

        params = {
            "jql": jql,
            "startAt": start_at,
            "maxResults": max_results,
            "fields": "key,summary,status,assignee,priority,updated" 
        }
        
        response = requests.get(
            url,
            headers=headers,
            params=params
        )
        
        if response.status_code != 200:
            print(f"请求失败！状态码: {response.status_code}")
            # print(response.text)
            break

        data = response.json()
        issues = data.get("issues", [])
        all_issues.extend(issues)
        
        # 如果已获取全部结果，退出循环
        if len(issues) < max_results:
            break
            
        start_at += max_results
    
    return all_issues


def get_tmx_info(issues):
    all_tmx_info = {}
    headers = {
        "Authorization": f"Bearer {API_TOKEN}",
        "Content-Type": "application/json"
    }

    for issue in issues:
        summary = issue["fields"]["summary"]
        # 解析数据
        version_type = summary.split("_")[0]
        
        if version_type == "STAR3.0":
            jira_key = issue["key"]
            response = requests.get(
                f"{JIRA_URL}/rest/api/2/issue/{jira_key}",
                headers=headers
            )
            data = response.json()
            passed = failed = nt = 0
            # 标记 BLOCKED 的值
            blocked_found = False
            for i,d in data.items():
                if type(d) == dict:
                    # 得到 test 的所有状态
                    custome = d.get("customfield_11120")
                    total = custome.get("count")
                    status = custome.get("statuses")
                    for i in status:
                        if i.get("name") == "PASS":
                            passed = i.get("statusCount")
                        elif i.get("name") == "FAIL":
                            failed = i.get("statusCount")
                        elif i.get("name") == "BLOCKED":
                            nt = i.get("statusCount")
                            if nt != 0:
                                nt = i.get("statusCount")
                                # 标记已找到有效的 BLOCKED
                                blocked_found = True
                        # 如果 blocked 没有值，取TODO的值
                        elif i.get("name") == "TODO" and not blocked_found:
                            nt = i.get("statusCount")
                    all_tmx_info[jira_key]={
                        "total": total,
                        "passed": passed,
                        "failed": failed,
                        "nt": nt
                    }
        elif version_type == "STAR3.5":
            jira_key = issue["key"]
            response = requests.get(
                f"{JIRA_URL}/rest/api/2/issue/{jira_key}",
                headers=headers
            )
            data = response.json()
            passed = failed = nt = 0
            # 标记 BLOCKED 的值
            blocked_found = False
            for i,d in data.items():
                if type(d) == dict:
                    # 得到 test 的所有状态
                    custome = d.get("customfield_11120")
                    total = custome.get("count")
                    status = custome.get("statuses")
                    for i in status:
                        if i.get("name") == "PASS":
                            passed = i.get("statusCount")
                        elif i.get("name") == "FAIL":
                            failed = i.get("statusCount")
                        elif i.get("name") == "BLOCKED":
                            nt = i.get("statusCount")
                            if nt != 0:
                                nt = i.get("statusCount")
                                # 标记已找到有效的 BLOCKED
                                blocked_found = True
                        # 如果 blocked 没有值，取TODO的值
                        elif i.get("name") == "TODO" and not blocked_found:
                            nt = i.get("statusCount")

                    all_tmx_info[jira_key]={
                        "total": total,
                        "passed": passed,
                        "failed": failed,
                        "nt": nt
                    }
    return all_tmx_info
        

def query_mr(mr_label,GITLAB_URL,GROUP_PATH,ACCESS_TOKEN):
    all_mrs=[]
    new_label = f"integrated_{mr_label}"
    page = 1
    per_page = 100  # 每页最大数量
    
    # 构建API请求URL
    url = f"{GITLAB_URL}/api/v4/groups/{GROUP_PATH}/merge_requests"
    # 请求头
    headers = {
        "PRIVATE-TOKEN": ACCESS_TOKEN
    } 
    # 请求参数
    params = {
        "scope": "all",
        "page": page,
        "per_page": per_page,
        "labels": new_label
    }
        
    # 发送GET请求
    response = requests.get(url, params=params, headers=headers)       
    if response.status_code != 200:
        print(f"请求失败，状态码: {response.status_code}")
        
    mrs = response.json()
    
    for mr in mrs:
        mr_iid = mr.get("iid")
        all_mrs.append(mr_iid)
    return all_mrs


def write_excel(issues,tmx_info,file_path,workbook):   
    # 创建或获取工作表
    if "Star30" not in workbook.sheetnames:
        ws1 = workbook.create_sheet("Star30")
    else:
        ws1 = workbook["Star30"]
        
    if "Star35" not in workbook.sheetnames:
        ws2 = workbook.create_sheet("Star35")
    else:
        ws2 = workbook["Star35"]

    star30_row_headers = ["i3 RSU Star 3.0", "Passed", "Blocked", "Failed", "NT", "Total", "MR Number"]
    star35_row_headers = ["i3 RSU Star 3.5", "Passed", "Blocked", "Failed", "NT", "Total", "MR Number"]
    # ws.append(row_headers)

    # 写入数据（第一列为行标头）
    for header in star30_row_headers:
        ws1.append([header])
    for header in star35_row_headers:
        ws2.append([header])

    index = 1
    count = 1
    for issue in issues:
        for tmx,data in tmx_info.items():
            key = issue["key"]
            summary = issue["fields"]["summary"]
            
            # 解析数据
            version_type = summary.split("_")[0]
            build_date = summary.split("_")[-1]
            build_type = "_".join(summary.split("_")[1:3])
            mr_label = summary.split("_")[1]

            # 写入第二列（B 列）的第一行
            
            if version_type == "STAR3.0":
                if key == tmx:
                    index = index + 1 
                    mr_number = len(query_mr(mr_label,GITLAB_URL,GROUP_PATH,ACCESS_TOKEN))
                    
                    # print(tmx,data.get("passed"),data.get("failed"),data.get("nt"))
                    cell_value = f"{build_date}\n{build_type}"
                    ws1.cell(row=1, column=index, value=cell_value)
                    # 设置单元格自动换行
                    ws1.cell(row=1, column=index).alignment = Alignment(wrap_text=True)
                    ws1.cell(row=2, column=index, value=data.get("passed"))
                    ws1.cell(row=3, column=index, value=0)
                    ws1.cell(row=4, column=index, value=data.get("failed"))
                    ws1.cell(row=5, column=index, value=data.get("nt"))
                    ws1.cell(row=6, column=index, value=data.get("total"))
                    ws1.cell(row=7, column=index, value=mr_number)
            elif version_type == "STAR3.5":
                if key == tmx:
                    count = count + 1 
                    mr_number = len(query_mr(mr_label,GITLAB_URL,GROUP_PATH,ACCESS_TOKEN))
                    
                    # print(tmx,data.get("passed"),data.get("failed"),data.get("nt"))
                    cell_value = f"{build_date}\n{build_type}"
                    ws2.cell(row=1, column=count, value=cell_value)
                    # 设置单元格自动换行
                    ws2.cell(row=1, column=count).alignment = Alignment(wrap_text=True)
                    ws2.cell(row=2, column=count, value=data.get("passed"))
                    ws2.cell(row=3, column=count, value=0)
                    ws2.cell(row=4, column=count, value=data.get("failed"))
                    ws2.cell(row=5, column=count, value=data.get("nt"))
                    ws2.cell(row=6, column=count, value=data.get("total"))
                    ws2.cell(row=7, column=count, value=mr_number)

    # 保存 Excel 文件
    workbook.save("rsu_staging_trend.xlsx")
    print("Excel 文件已生成：rsu_staging_trend.xlsx")

# 读取excel表格的内容，并保持原始内容不变更
def safe_read_excel(file_path):
    if not os.path.exists(file_path):
        print(f"错误: 文件不存在 - {file_path}")
        return None
    
    try:
        workbook = load_workbook(file_path, data_only=False)
        return workbook
    except Exception as e:
        print(f"错误: 无法读取Excel文件 - {e}")
        return None


if __name__ == '__main__':
    config = ConfigParser()
    config.read("/home/gaoyuxia/config.ini")
    ACCESS_TOKEN = config.get("gitlab","token")
    GITLAB_URL = config.get("gitlab","url")
    # JIRA info
    JIRA_URL = config.get("jira","JIRA_URL")
    API_TOKEN = config.get("jira","API_TOKEN")
    GROUP_PATH = "RSE"

    file_path = "RSU_staging_Trend.xlsx"

    issues = get_jira_issues(jql_query)
    # print(f"共找到 {len(issues)} 个问题!")
    tmx_info = get_tmx_info(issues)
    # 先读取整个 excel 表格的内容容
    workbook = safe_read_excel(file_path)
    # 将读取的内容操作持不变
    workbook.save("RSU_staging_Trend.xlsx")
    write_excel(issues,tmx_info,file_path,workbook)